#!/usr/bin/env python3
"""Generate example/ACME_Integrated_Financials.xlsx — a miniature, fully consistent
demonstration of the BAV workbook architecture (fictional company, 3 fiscal years,
one model scenario). Every pattern matches the real pipeline: hardcoded source tabs
with date headers, a formula-linked Condensed tab driven by an interactive
classification table (dropdowns), ALT DuPont, and a residual-income Model tab.

The numbers are constructed so all integrity checks tie exactly:
  - Balance sheet: retained earnings is the plug -> A = L + E every year
  - Cash flow: derived from balance-sheet deltas + income statement -> sums to Δcash
  - Condensed: implied equity (NOA - Net Debt) equals reported equity exactly
Run: python3 build_sample.py   (asserts all ties before writing)
"""
import datetime, os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
YEARS = [2023, 2024, 2025]
COLS = ['B', 'C', 'D']
BLUE = Font(color='0000FF'); BOLD = Font(bold=True); ITAL = Font(italic=True, size=9)
YEL = PatternFill('solid', start_color='FFF2CC')
GRN = PatternFill('solid', start_color='D9EAD3')
GRY = PatternFill('solid', start_color='F2F2F2')
HDR = PatternFill('solid', start_color='D9E2F3')
INP = PatternFill('solid', start_color='E8F0FE')
NUM = '#,##0;(#,##0)'

# ---------------- data (USD millions) ----------------
IS = {  # sign convention: costs negative
    'Revenues':                [1000, 1150, 1322],
    'Cost of revenues':        [-400, -460, -529],
    'Research and development':[-150, -172, -198],
    'Selling, general and administrative': [-200, -230, -264],
    'Interest expense':        [-10, -10, -10],
    'Interest income':         [5, 5, 5],
}
BS_ASSETS = {
    'Cash and cash equivalents':  [100, 130, 168],
    'Short-term investments':     [50, 55, 60],
    'Accounts receivable, net':   [120, 138, 159],
    'Inventory':                  [80, 92, 106],
    'Property and equipment, net':[400, 460, 529],
    'Goodwill':                   [60, 60, 60],
    'Other non-current assets':   [30, 33, 36],
}
BS_LIABS = {
    'Accounts payable':                 [70, 80, 92],
    'Accrued compensation':             [40, 46, 53],
    'Deferred revenue':                 [30, 35, 40],
    'Short-term debt':                  [20, 20, 20],
    'Long-term debt':                   [200, 210, 220],
    'Operating lease liabilities':      [40, 44, 48],
    'Other non-current liabilities':    [25, 28, 31],
}
CS = [150, 150, 150]
DA = [None, 70, 80]            # depreciation & amortization (CF years only)
TAXRATE = 0.21

CLASSIFY = {                    # textbook defaults — your _conventions.md overrides these
    'Cash and cash equivalents': 'Financial Asset',
    'Short-term investments': 'Financial Asset',
    'Accounts receivable, net': 'Operating Working Capital Asset',
    'Inventory': 'Operating Working Capital Asset',
    'Property and equipment, net': 'Operating Long-Term Asset',
    'Goodwill': 'Operating Long-Term Asset',
    'Other non-current assets': 'Operating Long-Term Asset',
    'Accounts payable': 'Operating Working Capital Liability',
    'Accrued compensation': 'Operating Working Capital Liability',
    'Deferred revenue': 'Operating Working Capital Liability',
    'Short-term debt': 'Financial Liability',
    'Long-term debt': 'Financial Liability',
    'Operating lease liabilities': 'Operating Long-Term Liability',
    'Other non-current liabilities': 'Operating Long-Term Liability',
    'Common stock and paid-in capital': 'Equity',
    'Retained earnings': 'Equity',
}
CATEGORIES = ('Operating Working Capital Asset,Operating Working Capital Liability,'
              'Operating Long-Term Asset,Operating Long-Term Liability,'
              'Financial Asset,Financial Liability,Equity,Exclude')

# Model_Base assumptions (10-year vectors; ratio vectors: [0] documents the anchor)
GROWTH = [0.12, 0.11, 0.10, 0.09, 0.08, 0.07, 0.06, 0.05, 0.045, 0.04]
MARGIN = [0.20, 0.195, 0.19, 0.18, 0.17, 0.16, 0.155, 0.15, 0.145, 0.14]
NOWC_V = [0.061, 0.06, 0.06, 0.06, 0.06, 0.06, 0.06, 0.06, 0.06, 0.06]
NOLA_V = [0.413, 0.44, 0.46, 0.48, 0.50, 0.52, 0.54, 0.56, 0.58, 0.60]
BETA, RF, ERP, GTERM, SHARES = 1.1, 0.04, 0.05, 0.03, 100

# ---------------- derive + assert consistency ----------------
ni, pretax, tax = [], [], []
for i in range(3):
    op = sum(IS[k][i] for k in list(IS)[:4])
    pt = op + IS['Interest expense'][i] + IS['Interest income'][i]
    tx = -round(pt * TAXRATE)
    pretax.append(pt); tax.append(tx); ni.append(pt + tx)

assets = [sum(v[i] for v in BS_ASSETS.values()) for i in range(3)]
liabs = [sum(v[i] for v in BS_LIABS.values()) for i in range(3)]
re = [assets[i] - liabs[i] - CS[i] for i in range(3)]          # equity plug
div = [None] + [re[i-1] + ni[i] - re[i] for i in (1, 2)]

cf = {'years': YEARS[1:], 'CFO': [], 'CapEx': [], 'CFI': [], 'CFF': [], 'dCash': []}
for i in (1, 2):
    d = lambda key, src: src[key][i] - src[key][i-1]
    dnwc = (d('Accounts receivable, net', BS_ASSETS) + d('Inventory', BS_ASSETS)
            - d('Accounts payable', BS_LIABS) - d('Accrued compensation', BS_LIABS)
            - d('Deferred revenue', BS_LIABS))
    cfo = (ni[i] + DA[i] - dnwc + d('Operating lease liabilities', BS_LIABS)
           + d('Other non-current liabilities', BS_LIABS))
    capex = d('Property and equipment, net', BS_ASSETS) + DA[i]
    cfi = -(capex + d('Short-term investments', BS_ASSETS)
            + d('Other non-current assets', BS_ASSETS) + d('Goodwill', BS_ASSETS))
    cff = (d('Short-term debt', BS_LIABS) + d('Long-term debt', BS_LIABS) - div[i])
    dcash = d('Cash and cash equivalents', BS_ASSETS)
    assert cfo + cfi + cff == dcash, f'CF does not tie in {YEARS[i]}'
    for k, v in zip(('CFO', 'CapEx', 'CFI', 'CFF', 'dCash'), (cfo, capex, cfi, cff, dcash)):
        cf[k].append(v)
for i in range(3):
    assert assets[i] == liabs[i] + CS[i] + re[i]
print('checksums: BS ties, CF ties, all years')

# ---------------- workbook ----------------
wb = openpyxl.Workbook()
wb.remove(wb.active)

def header(ws, title, sub):
    ws['A1'] = f'ACME Corporation (ACME) — {title}'; ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = sub; ws['A2'].font = ITAL
    ws['A3'] = 'Sample data — fictional company for demonstration'
    ws['A3'].font = ITAL
    ws.column_dimensions['A'].width = 42
    for c in COLS + ['E', 'F']:
        ws.column_dimensions[c].width = 14

def yearrow(ws, row=6, ncols=3):
    ws[f'A{row}'] = 'Line Item'; ws[f'A{row}'].font = BOLD
    for j in range(ncols):
        c = ws[f'{COLS[j]}{row}']
        c.value = datetime.date(YEARS[j + (3 - ncols)], 12, 31)
        c.number_format = 'MMM DD, YYYY'; c.font = BOLD

# Income Statement
ws = wb.create_sheet('Income Statement')
header(ws, 'Income Statement', 'Units: USD millions · only the source tabs hold hardcoded data')
yearrow(ws)
rows = [('Revenues', IS['Revenues']), ('Cost of revenues', IS['Cost of revenues']),
        ('Gross profit', None), ('Research and development', IS['Research and development']),
        ('Selling, general and administrative', IS['Selling, general and administrative']),
        ('Income from operations', None), ('Interest expense', IS['Interest expense']),
        ('Interest income', IS['Interest income']), ('Income before income taxes', None),
        ('Provision for income taxes', tax), ('Net income', None)]
for r, (label, vals) in enumerate(rows, start=7):
    ws[f'A{r}'] = label
    for j, c in enumerate(COLS):
        cell = ws[f'{c}{r}']; cell.number_format = NUM
        if label == 'Gross profit':
            cell.value = f'={c}7+{c}8'
        elif label == 'Income from operations':
            cell.value = f'={c}9+{c}10+{c}11'
        elif label == 'Income before income taxes':
            cell.value = f'={c}12+{c}13+{c}14'
        elif label == 'Net income':
            cell.value = f'={c}15+{c}16'
        else:
            cell.value = vals[j]
        if vals is None:
            cell.font = BOLD
IS_ROW = {'rev': 7, 'pretax': 15, 'tax': 16, 'ni': 17, 'intexp': 13, 'intinc': 14}

# Balance Sheet
ws = wb.create_sheet('Balance Sheet')
header(ws, 'Balance Sheet', 'Units: USD millions')
yearrow(ws)
r = 7; BS_ROW = {}
def put_section(ws, title, items, r):
    ws[f'A{r}'] = title; ws[f'A{r}'].font = BOLD; r += 1
    for label, vals in items:
        ws[f'A{r}'] = label; BS_ROW[label] = r
        for j, c in enumerate(COLS):
            ws[f'{c}{r}'] = vals[j]; ws[f'{c}{r}'].number_format = NUM
        r += 1
    return r
r = put_section(ws, 'ASSETS', list(BS_ASSETS.items()), r)
ws[f'A{r}'] = 'Total assets'; ws[f'A{r}'].font = BOLD
for j, c in enumerate(COLS):
    ws[f'{c}{r}'] = f'=SUM({c}8:{c}{r-1})'; ws[f'{c}{r}'].number_format = NUM; ws[f'{c}{r}'].font = BOLD
r += 2
r0 = r
r = put_section(ws, 'LIABILITIES', list(BS_LIABS.items()), r)
ws[f'A{r}'] = 'Total liabilities'; ws[f'A{r}'].font = BOLD
for j, c in enumerate(COLS):
    ws[f'{c}{r}'] = f'=SUM({c}{r0+1}:{c}{r-1})'; ws[f'{c}{r}'].number_format = NUM; ws[f'{c}{r}'].font = BOLD
r += 2
eq_items = [('Common stock and paid-in capital', CS), ('Retained earnings', re)]
r1 = r
r = put_section(ws, "STOCKHOLDERS' EQUITY", eq_items, r)
ws[f'A{r}'] = "Total stockholders' equity"; ws[f'A{r}'].font = BOLD
EQ_TOTAL_ROW = r
for j, c in enumerate(COLS):
    ws[f'{c}{r}'] = f'=SUM({c}{r1+1}:{c}{r-1})'; ws[f'{c}{r}'].number_format = NUM; ws[f'{c}{r}'].font = BOLD

# Cash Flow (derived years only)
ws = wb.create_sheet('Cash Flow')
header(ws, 'Cash Flow Statement', 'Units: USD millions · FY2023 omitted (no prior-year balance in sample)')
yearrow(ws, ncols=2)
cfrows = [('Net income', [ni[1], ni[2]]), ('Depreciation & amortization', DA[1:]),
          ('Working capital & other', [cf['CFO'][k] - ni[1+k] - DA[1+k] for k in (0, 1)]),
          ('Cash from operations', cf['CFO']), ('Capital expenditures', [-x for x in cf['CapEx']]),
          ('Investments & other', [cf['CFI'][k] + cf['CapEx'][k] for k in (0, 1)]),
          ('Cash from investing', cf['CFI']), ('Debt issued, net', [10, 10]),
          ('Dividends paid', [-div[1], -div[2]]), ('Cash from financing', cf['CFF']),
          ('Net change in cash', cf['dCash'])]
for r, (label, vals) in enumerate(cfrows, start=7):
    ws[f'A{r}'] = label
    if label.startswith('Cash from') or label.startswith('Net change'):
        ws[f'A{r}'].font = BOLD
    for j, c in enumerate(COLS[:2]):
        ws[f'{c}{r}'] = vals[j]; ws[f'{c}{r}'].number_format = NUM

# Condensed Financials — the interactive heart
ws = wb.create_sheet('Condensed Financials')
header(ws, 'Condensed Financials',
       'All aggregates are SUMIFs over the CLASSIFICATION TABLE below — toggle column E to reclassify')
yearrow(ws, row=5)
ws['A7'] = 'A. CONDENSED INCOME STATEMENT'; ws['A7'].font = BOLD; ws['A7'].fill = GRY
nopat_rows = [('Net income', lambda c: f"='Income Statement'!{c}{IS_ROW['ni']}"),
              ('Interest expense (abs)', lambda c: f"=ABS('Income Statement'!{c}{IS_ROW['intexp']})"),
              ('Interest income', lambda c: f"='Income Statement'!{c}{IS_ROW['intinc']}"),
              ('Net interest expense', lambda c: f'={c}9-{c}10'),
              ('Effective tax rate', lambda c: f"=IFERROR(ABS('Income Statement'!{c}{IS_ROW['tax']})/'Income Statement'!{c}{IS_ROW['pretax']},0)"),
              ('Net interest after tax', lambda c: f'={c}11*(1-{c}12)'),
              ('NOPAT', lambda c: f'={c}8+{c}13')]
for r, (label, f) in enumerate(nopat_rows, start=8):
    ws[f'A{r}'] = label
    for c in COLS:
        ws[f'{c}{r}'] = f(c)
        ws[f'{c}{r}'].number_format = '0.0%' if 'rate' in label else NUM
ws['A14'].font = BOLD
for c in COLS: ws[f'{c}14'].fill = YEL

TBL0, TBLN = 33, 33 + len(CLASSIFY) - 1
agg = [('Operating Working Capital Assets', 'Operating Working Capital Asset', 17),
       ('Operating Working Capital Liabilities', 'Operating Working Capital Liability', 18),
       ('NET OPERATING WORKING CAPITAL (NOWC)', None, 19),
       ('Operating Long-Term Assets', 'Operating Long-Term Asset', 21),
       ('Operating Long-Term Liabilities', 'Operating Long-Term Liability', 22),
       ('NET OPERATING LT ASSETS (NOLA)', None, 23),
       ('NET OPERATING ASSETS (NOA)', None, 25),
       ('Financial Assets', 'Financial Asset', 27),
       ('Financial Liabilities', 'Financial Liability', 28),
       ('NET DEBT', None, 29),
       ('EQUITY (NOA − Net Debt)', None, 30),
       ('Reported equity', None, 31)]
ws['A16'] = 'B. CONDENSED BALANCE SHEET'; ws['A16'].font = BOLD; ws['A16'].fill = GRY
for label, cat, r in agg:
    ws[f'A{r}'] = label; ws[f'A{r}'].font = BOLD
    for c in COLS:
        cell = ws[f'{c}{r}']; cell.number_format = NUM
        if cat:
            cell.value = f'=SUMIF($E${TBL0}:$E${TBLN},"{cat}",{c}${TBL0}:{c}${TBLN})'
        elif 'NOWC' in label: cell.value = f'={c}17-{c}18'; cell.fill = YEL
        elif 'NOLA' in label: cell.value = f'={c}21-{c}22'; cell.fill = YEL
        elif '(NOA)' in label: cell.value = f'={c}19+{c}23'; cell.fill = GRN
        elif 'NET DEBT' in label: cell.value = f'={c}28-{c}27'; cell.fill = GRN
        elif 'EQUITY (NOA' in label: cell.value = f'={c}25-{c}29'; cell.fill = GRN
        else: cell.value = f"='Balance Sheet'!{c}{EQ_TOTAL_ROW}"
ws['A32'] = 'C. CLASSIFICATION TABLE — click any cell in column E for the dropdown'
ws['A32'].font = BOLD; ws['A32'].fill = GRY
for i, (label, cat) in enumerate(CLASSIFY.items()):
    r = TBL0 + i
    ws[f'A{r}'] = label
    src = BS_ROW.get(label)
    for c in COLS:
        ws[f'{c}{r}'] = f"='Balance Sheet'!{c}{src}"; ws[f'{c}{r}'].number_format = NUM
    e = ws[f'E{r}']; e.value = cat; e.font = BLUE; e.fill = INP
dv = DataValidation(type='list', formula1=f'"{CATEGORIES}"', allow_blank=False, showDropDown=False)
dv.add(f'E{TBL0}:E{TBLN}')
ws.add_data_validation(dv)
ws.column_dimensions['E'].width = 34

# ALT DuPont
ws = wb.create_sheet('ALT DuPont')
header(ws, 'ALT DuPont', 'ROE = RNOA + FLEV × Spread · all inputs formula-linked to Condensed Financials')
yearrow(ws, row=5)
dup = [('NOPAT margin', lambda c: f"=IFERROR('Condensed Financials'!{c}14/'Income Statement'!{c}7,0)"),
       ('Avg NOA', lambda c: None), ('Asset turnover', lambda c: None),
       ('RNOA', lambda c: None), ('Avg net debt', lambda c: None),
       ('After-tax cost of debt', lambda c: None), ('Spread', lambda c: None),
       ('Avg equity', lambda c: None), ('FLEV', lambda c: None),
       ('ROE (RNOA + FLEV×Spread)', lambda c: None),
       ('Actual ROE (NI / Avg equity)', lambda c: None)]
for r, (label, _) in enumerate(dup, start=7):
    ws[f'A{r}'] = label
for j, c in enumerate(COLS):
    p = COLS[j-1] if j else None
    def setf(r, f, fmt='0.0%'):
        ws[f'{c}{r}'] = f; ws[f'{c}{r}'].number_format = fmt
    setf(7, f"=IFERROR('Condensed Financials'!{c}14/'Income Statement'!{c}7,0)")
    if p:
        setf(8, f"=('Condensed Financials'!{c}25+'Condensed Financials'!{p}25)/2", NUM)
        setf(9, f"=IFERROR('Income Statement'!{c}7/{c}8,0)", '0.00')
        setf(10, f'={c}7*{c}9'); ws[f'{c}10'].fill = YEL
        setf(11, f"=('Condensed Financials'!{c}29+'Condensed Financials'!{p}29)/2", NUM)
        setf(12, f"=IFERROR('Condensed Financials'!{c}13/{c}11,0)")
        setf(13, f'={c}10-{c}12')
        setf(14, f"=('Condensed Financials'!{c}30+'Condensed Financials'!{p}30)/2", NUM)
        setf(15, f'=IFERROR({c}11/{c}14,0)', '0.00')
        setf(16, f'={c}10+{c}13*{c}15'); ws[f'{c}16'].fill = YEL
        setf(17, f"=IFERROR('Condensed Financials'!{c}8/{c}14,0)")
    else:
        for r in range(8, 18): ws[f'{c}{r}'] = 'n/a'

# Model_Base (compact, faithful row layout)
ws = wb.create_sheet('Model_Base')
header(ws, 'Residual Income Model — BASE', '10-year forecast · blue cells are editable inputs')
md = [('Beta', BETA), ('Equity risk premium', ERP), ('Risk-free rate', RF)]
for r, (label, v) in enumerate(md, start=4):
    ws[f'A{r}'] = label; ws[f'B{r}'] = v; ws[f'B{r}'].font = BLUE
ws['A7'] = 'Cost of equity (Ke)'; ws['B7'] = '=B6+B4*B5'
FC = ['C','D','E','F','G','H','I','J','K','L']
ws['A9'] = 'Year'; ws['B9'] = YEARS[-1]
for i, c in enumerate(FC):
    ws[f'{c}9'] = YEARS[-1] + 1 + i; ws[f'{c}9'].font = BOLD
rows = [('Revenue growth', 10, GROWTH, '0.0%'), ('NOPAT margin', 11, MARGIN, '0.0%'),
        ('NOWC / sales', 12, NOWC_V, '0.0%'), ('NOLA / sales', 13, NOLA_V, '0.0%')]
for label, r, vec, fmt in rows:
    ws[f'A{r}'] = label
    for i, c in enumerate(FC):
        if r >= 12 and i == 0:
            ws[f'{c}{r}'] = f'={c}16/{c}15' if r == 12 else f'={c}17/{c}15'
        else:
            ws[f'{c}{r}'] = vec[i]; ws[f'{c}{r}'].font = BLUE
        ws[f'{c}{r}'].number_format = fmt
ws['A15'] = 'Sales'; ws['B15'] = f"='Income Statement'!D{IS_ROW['rev']}"
ws['A16'] = 'NOWC (beg)'; ws['C16'] = "='Condensed Financials'!D19"
ws['A17'] = 'NOLA (beg)'; ws['C17'] = "='Condensed Financials'!D23"
ws['A18'] = 'Net debt (beg)'; ws['C18'] = "='Condensed Financials'!D29"
ws['A19'] = 'Equity (beg)'; ws['C19'] = '=C16+C17-C18'
for i, c in enumerate(FC):
    p = 'B' if i == 0 else FC[i-1]
    ws[f'{c}15'] = f'={p}15*(1+{c}10)'
    if i > 0:
        ws[f'{c}16'] = f'={c}12*{c}15'; ws[f'{c}17'] = f'={c}13*{c}15'
        ws[f'{c}18'] = f'=$C$18/($C$16+$C$17)*({c}16+{c}17)'
        ws[f'{c}19'] = f'={c}16+{c}17-{c}18'
    for r in range(15, 20): ws[f'{c}{r}'].number_format = NUM
ws['A21'] = 'NOPAT'; ws['A22'] = 'Net income'; ws['A23'] = 'Abnormal earnings'
ws['A24'] = 'PV factor'; ws['A25'] = 'PV of AE'
for i, c in enumerate(FC):
    p = FC[i-1] if i else None
    ws[f'{c}21'] = f'={c}15*{c}11'
    ws[f'{c}22'] = f'={c}21-{c}18*0.03'        # simple 3% after-tax cost on net debt
    ws[f'{c}23'] = f'={c}22-{c}19*$B$7'; ws[f'{c}23'].fill = GRN
    ws[f'{c}24'] = '=1/(1+$B$7)' if i == 0 else f'={p}24/(1+$B$7)'
    ws[f'{c}25'] = f'={c}23*{c}24'
    for r in (21, 22, 23, 25): ws[f'{c}{r}'].number_format = NUM
    ws[f'{c}24'].number_format = '0.000'
ws['A27'] = 'Sum PV(AE)'; ws['C27'] = '=SUM(C25:L25)'
ws['A28'] = f'Terminal value (g={GTERM:.0%})'; ws['C28'] = f'=L23*(1+{GTERM})/($B$7-{GTERM})*L24'
ws['A29'] = 'Beginning book equity'; ws['C29'] = '=C19'
ws['A30'] = 'Intrinsic value of equity'; ws['C30'] = '=C27+C28+C29'; ws['C30'].font = BOLD
ws['A31'] = 'Diluted shares (M)'; ws['C31'] = SHARES; ws['C31'].font = BLUE
ws['A32'] = 'INTRINSIC VALUE PER SHARE'; ws['C32'] = '=C30/C31'
ws['A32'].font = BOLD; ws['C32'].font = BOLD; ws['C32'].fill = GRN
ws['C32'].number_format = '$#,##0.00'
for r in range(27, 31): ws[f'C{r}'].number_format = NUM
for c in FC: ws.column_dimensions[c].width = 11
ws.freeze_panes = 'B1'
ws['A34'] = ('Note: real builds carry Model_Bear / Model_Base / Model_Bull plus a '
             'Scenario_Summary with probability weighting — this sample shows one scenario.')
ws['A34'].font = ITAL

out = os.path.join(HERE, 'ACME_Integrated_Financials.xlsx')
wb.save(out)
print('wrote', out)
