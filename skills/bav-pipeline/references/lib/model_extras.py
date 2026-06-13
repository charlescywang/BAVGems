#!/usr/bin/env python3
"""Model-tab extras shared across tickers (the skill's canonical post-build pass):

  1. Terminal growth (B9) LINKED to the terminal-year revenue growth cell (U11),
     so terminal g always equals the last explicit forecast-year revenue growth
     rather than a stale hardcoded literal. (analyst pattern: link terminal g to the last explicit-period revenue growth.)
  2. A "Historical Average" column (W) on each Model tab: AVERAGE(C{r}:K{r}) of the
     historical (as-filed) span for the four key driver rows — revenue growth (11),
     NOPAT margin (12), NOWC/Sales (14), Net-Op-LT-Assets/Sales (15) — so the
     forecast vectors (L:U) can be eyeballed against the historical norm.

Idempotent and engine-agnostic (operates on the workbook). Apply after the model
build (or call from it). `repoint_u46_to_b9=True` also rewrites a terminal-value
formula that HARDCODES the growth rate to reference $B$9 instead — only safe when
the terminal revenue growth (U11) is the intended perpetuity growth (NOT the case
for a ticker that deliberately caps terminal g below its last explicit growth).
"""
# revenue growth, NOPAT margin, NOWC/Sales, NOLA/Sales, ROE (Return on Beg Equity),
# RNOA (Return on Beg NOA) — the key drivers AND the realized returns
HIST_AVG_ROWS = (11, 12, 14, 15, 36, 37)
HIST_COLS = 'C:K'                  # as-filed historical span (FY-9 .. FY anchor)


def apply_model_extras(wb, model_fmt='Model_{}', names=('Bear', 'Base', 'Bull'),
                       repoint_u46_to_b9=False):
    from openpyxl.styles import Font
    BLUE = Font(name='Calibri', size=11, color='FF0000FF')   # editable-input convention
    HDR = Font(name='Calibri', size=11, bold=True)
    changed = []
    for n in names:
        ws = wb[model_fmt.format(n)]
        # (1) terminal growth linked to terminal-year revenue growth
        ws['B9'] = '=U11'
        ws['B9'].font = BLUE
        if not ws['A9'].value:
            ws['A9'] = 'Terminal growth (g)'
            ws['A9'].font = Font(name='Calibri', size=11)
        if repoint_u46_to_b9 and ws['U46'].value and '$B$9' not in str(ws['U46'].value):
            ws['U46'] = '=U42*(1+$B$9)/($B$5-$B$9)'
            changed.append(f'{n}:U46->B9')
        # (2) historical-average column W
        ws['W10'] = 'Historical Average'
        ws['W10'].font = HDR
        for r in HIST_AVG_ROWS:
            c = ws[f'W{r}']
            c.value = f'=AVERAGE(C{r}:K{r})'
            c.number_format = ws[f'L{r}'].number_format  # match the row's format (%, ratio)
        changed.append(f'{n}:B9=U11,W')
    return changed
