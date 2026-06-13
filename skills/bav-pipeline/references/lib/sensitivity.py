#!/usr/bin/env python3
"""Two-dimensional valuation sensitivity grids for the Scenario_Summary tab.

A terminal-heavy residual-income valuation is, in order, most sensitive to the cost
of equity, then terminal profitability (margin / asset intensity), then terminal
growth (muted because the terminal is ABNORMAL earnings, not free cash flow). So we
grid IVPS against the two highest-impact pairs, centered on the Base case:

  Grid 1: IVPS vs Cost of Equity (rows) × Terminal growth g (cols)
  Grid 2: IVPS vs Cost of Equity (rows) × Terminal NOPAT margin (cols)

Each cell is a LIVE closed-form of the RIM (verified to reproduce the engine exactly):
operating NI_t and beginning equity_t are independent of Ke/g, so

  IVPS(Ke,g)  = (BV0 + Σ_t (NI_t − Ke·Eq_t)/(1+Ke)^t
                     + (NI_10 − Ke·Eq_10)(1+g)/(Ke−g)/(1+Ke)^10) / shares

and the margin grid replaces year-10 NI with sales_10·m − interest_10. Headers are
centered on Model_Base!B5 (Ke), B9 (g), U12 (terminal margin) via offsets, so the
grids recompute when the Base case changes; the center cell equals Model_Base!L51.
"""
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule

F_HEADER = Font(name='Calibri', size=11, bold=True)
F_LABEL = Font(name='Calibri', size=11, bold=False)
F_CAPTION = Font(name='Calibri', size=9, italic=True)
FILL_HDR = PatternFill('solid', start_color='FFD9E2F3', end_color='FFD9E2F3')
FILL_CTR = PatternFill('solid', start_color='FFD9EAD3', end_color='FFD9EAD3')   # base-case anchor
WRAP = Alignment(wrap_text=True, vertical='top')
NF_PS = '#,##0.00'
KE_OFFSETS = (-0.02, -0.01, 0.0, 0.01, 0.02)        # rows: Ke ±2pp
G_OFFSETS = (-0.01, -0.005, 0.0, 0.005, 0.01)       # cols: terminal g ±1pp
M_OFFSETS = (-0.03, -0.015, 0.0, 0.015, 0.03)       # cols: terminal margin ±3pp
COLS = ('B', 'C', 'D', 'E', 'F')                    # 5 value columns
EXP10 = '{1,2,3,4,5,6,7,8,9,10}'
EXP9 = '{1,2,3,4,5,6,7,8,9}'


def _ke_g_cell(M, ke_ref, g_ref):
    return (f'=(({M}!$L$48'
            f'+SUMPRODUCT(({M}!$L$40:$U$40-{ke_ref}*{M}!$L$27:$U$27)/(1+{ke_ref})^{EXP10})'
            f'+({M}!$U$40-{ke_ref}*{M}!$U$27)*(1+{g_ref})/({ke_ref}-{g_ref})/(1+{ke_ref})^10)'
            f'/{M}!$L$50)')


def _ke_m_cell(M, ke_ref, m_ref):
    # year-10 NI rebuilt from a varied terminal margin: sales_10·m − interest_10
    ni10 = f'({M}!$U$31*{m_ref}-{M}!$U$33-{ke_ref}*{M}!$U$27)'
    return (f'=(({M}!$L$48'
            f'+SUMPRODUCT(({M}!$L$40:$T$40-{ke_ref}*{M}!$L$27:$T$27)/(1+{ke_ref})^{EXP9})'
            f'+{ni10}/(1+{ke_ref})^10'
            f'+{ni10}*(1+{M}!$B$9)/({ke_ref}-{M}!$B$9)/(1+{ke_ref})^10)'
            f'/{M}!$L$50)')


def _grid(ss, top, title, col_center_ref, col_offsets, col_fmt, cell_fn, M):
    ss.cell(row=top, column=1, value=title).font = F_HEADER
    ss.cell(row=top + 1, column=1, value='Ke ↓ / param →').font = F_CAPTION
    # column headers (param values), centered on col_center_ref
    for j, off in enumerate(col_offsets):
        c = ss[f'{COLS[j]}{top+1}']
        c.value = f'={col_center_ref}{"+" if off>=0 else "-"}{abs(off)}'
        c.number_format = col_fmt
        c.font = F_HEADER
        c.fill = FILL_HDR
    # row headers (Ke values) + cells
    for i, koff in enumerate(KE_OFFSETS):
        r = top + 2 + i
        kref = f'$A${r}'
        kc = ss[f'A{r}']
        kc.value = f'=Model_{{}}!$B$5{"+" if koff>=0 else "-"}{abs(koff)}'.format('Base')
        kc.number_format = '0.00%'
        kc.font = F_HEADER
        kc.fill = FILL_HDR
        for j in range(5):
            cref = f'{COLS[j]}${top+1}'
            cell = ss[f'{COLS[j]}{r}']
            cell.value = cell_fn(M, kref, cref)
            cell.number_format = NF_PS
            if koff == 0.0 and col_offsets[j] == 0.0:    # base-case anchor cell (bold outline)
                cell.font = Font(name='Calibri', size=11, bold=True)
    # heatmap: red where IVPS is below the grid / baseline, green where above
    rng = f'{COLS[0]}{top+2}:{COLS[-1]}{top+1+len(KE_OFFSETS)}'
    ss.conditional_formatting.add(rng, ColorScaleRule(
        start_type='percentile', start_value=0, start_color='FFF8696B',     # red (low)
        mid_type='percentile', mid_value=50, mid_color='FFFFEB84',          # yellow (≈baseline)
        end_type='percentile', end_value=100, end_color='FF63BE7B'))        # green (high)
    return top + 2 + len(KE_OFFSETS)


def build_sensitivity_grids(ss, start_row, model_tab='Model_Base'):
    M = model_tab
    r = _grid(ss, start_row,
              'SENSITIVITY — IVPS vs Cost of Equity (rows) × Terminal Growth (cols), Base case',
              f'{M}!$B$9', G_OFFSETS, '0.00%', _ke_g_cell, M)
    r += 1
    r = _grid(ss, r,
              'SENSITIVITY — IVPS vs Cost of Equity (rows) × Terminal NOPAT Margin (cols), Base case',
              f'{M}!$U$12', M_OFFSETS, '0.0%', _ke_m_cell, M)
    note = ss.cell(row=r + 1, column=1, value=(
        'Live grids: each cell recomputes the Base-case IVPS at that Ke/parameter pair '
        '(green = current Base, equals Scenario_Summary IVPS). Valuation is most sensitive '
        'to Ke, then terminal margin/asset-intensity; terminal growth is muted (RIM terminal '
        'is abnormal earnings, not free cash flow).'))
    note.font = F_CAPTION
    ss.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=6)
    note.alignment = WRAP
    return r + 1
