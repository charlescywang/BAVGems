#!/usr/bin/env python3
"""Implied cost of capital (ICC) — engine-agnostic core for the BAV coverage system.

ICC = the discount rate r at which a scenario's residual-income IVPS equals the
market price. It is the implied expected return of buying at that price under that
scenario's forecast — the inverse of the valuation. Expected excess return
(ICC - Ke) is the universe ranking statistic (panel item 9): positive = the price
implies a return above our hurdle (undervalued / lean-buy), negative = below it
(overvalued / avoid).

This module is the canonical math + the one-price substrate (marks.jsonl). It is
ENGINE-AGNOSTIC: callers pass `ivps_fn(scenario_name, r) -> ivps`, a closure over
their ticker's own build_model (tickers can have different run_scenario
signatures, so the closure lives in each ticker's scripts/compute_icc.py shim).

Design notes that are load-bearing, not stylistic:
  - Bisection searches BOTH directions around Ke. IVPS(r) is monotone DECREASING in
    r, so if price < IVPS(Ke) the solving rate is ABOVE Ke (the old axis-D search
    only lowered Ke and so returned null when price < IVPS(Ke) — a bug, not a degeneracy).
  - Two true no-solution cases, reported as explicit SIGNALS not blanks:
      overvalued_past_floor : price > IVPS even at the terminal-growth floor
                              (no required return above perpetual growth justifies
                              the price) -> strong avoid.
      deep_cheap            : price < IVPS even at the 60% cap (model exceeds price
                              even at a punitive discount) -> rare, very cheap.
  - The probability-weighted ICC is only meaningful if the scenario IVPS ordering is
    monotone (Bear <= Base <= Bull). A non-monotone set (e.g. Base > Bull)
    averages contradictory per-scenario signals; we SUPPRESS the rank and flag
    "scenario set malformed" rather than print a confident-but-meaningless number.

ICC is a root-find, not a closed form, so any ICC written into a workbook/dashboard
is a COMPUTED value stamped with its as-of price — never a live formula. A stale
price silently mis-ranks the whole universe, so marks carry their date and callers
must surface staleness loudly.
"""
import json
import os

FLOOR_PAD = 0.002      # ICC lower bound = terminalGrowth + this (RIM diverges at r=g)
RATE_CAP = 0.60        # ICC upper bound; above this is economically meaningless
NEAR_FLOOR_BAND = 0.02 # ICC within this of terminalGrowth -> flag "no real compensation"
SANITY_MOVE = 0.50     # reject a new mark that moves >50% vs the last good one


# --------------------------------------------------------------------------- ICC

def solve_icc(ivps_at_r, price, term_g, cap=RATE_CAP):
    """Solve ivps_at_r(r) == price for r over (term_g+FLOOR_PAD, cap).

    ivps_at_r: callable r -> ivps (a closure that rebuilds the scenario at Ke=r).
    Returns (icc_or_None, status). status in {ok, overvalued_past_floor, deep_cheap}.
    """
    lo, hi = term_g + FLOOR_PAD, cap
    f = lambda r: ivps_at_r(r) - price
    flo, fhi = f(lo), f(hi)
    # f is decreasing in r: f(lo) is the most positive, f(hi) the most negative.
    if flo < 0:
        # even at the floor rate the model is below price -> price too high to justify
        return None, 'overvalued_past_floor'
    if fhi > 0:
        # even at the cap rate the model exceeds price -> price deeply below model
        return None, 'deep_cheap'
    for _ in range(100):
        mid = (lo + hi) / 2
        fm = f(mid)
        if abs(fm) < 1e-4:
            return mid, 'ok'
        if flo * fm < 0:
            hi = mid
        else:
            lo, flo = mid, fm
    return (lo + hi) / 2, 'ok'


def compute_icc(scenarios, ivps_fn, price, names=('Bear', 'Base', 'Bull')):
    """Full ICC result for one ticker at one price.

    scenarios: the assumptions['scenarios'] dict (each has probability, costOfEquity,
               terminalGrowth).
    ivps_fn:   callable (scenario_name, r) -> ivps, closing over the ticker engine.
    Returns the structure persisted to icc.json (see module docstring / shims).
    """
    out = {'price': round(price, 2), 'scenarios': {}, 'flags': []}
    iv_at_ke = {}
    for k in names:
        s = scenarios[k]
        ke, g, p = s['costOfEquity'], s['terminalGrowth'], s.get('probability')
        iv = ivps_fn(k, ke)
        iv_at_ke[k] = iv
        icc, status = solve_icc(lambda r: ivps_fn(k, r), price, g)
        excess = (icc - ke) if icc is not None else None
        near_floor = icc is not None and (icc - g) < NEAR_FLOOR_BAND
        out['scenarios'][k] = {
            'probability': p,
            'ivps': round(iv, 2),
            'ke': round(ke, 4),
            'term_g': round(g, 4),
            'icc': None if icc is None else round(icc, 4),
            'excess': None if excess is None else round(excess, 4),
            'status': status,
            'near_floor': near_floor,
        }

    # probability-weighted view
    probs = {k: scenarios[k].get('probability', 0) for k in names}
    psum = sum(probs.values()) or 1.0
    w_ke = sum(probs[k] / psum * scenarios[k]['costOfEquity'] for k in names)
    out['weighted'] = {'ke': round(w_ke, 4), 'ivps': round(
        sum(probs[k] / psum * iv_at_ke[k] for k in names), 2)}

    # monotonicity of the scenario hull (Bear <= Base <= Bull in IVPS)
    ordered = [iv_at_ke[k] for k in names]
    monotone = all(ordered[i] <= ordered[i + 1] + 1e-9 for i in range(len(ordered) - 1))
    out['monotonic'] = monotone
    out['max_scenario_ivps'] = round(max(iv_at_ke.values()), 2)
    out['prob_mix_feasible'] = price <= max(iv_at_ke.values()) + 1e-9

    all_ok = all(out['scenarios'][k]['status'] == 'ok' for k in names)
    if not monotone:
        # find the inversion to name it
        for i in range(len(names) - 1):
            if ordered[i] > ordered[i + 1] + 1e-9:
                out['flags'].append(
                    f'scenario set malformed: {names[i+1]} IVPS {ordered[i+1]:.2f} '
                    f'< {names[i]} IVPS {ordered[i]:.2f} — re-specify before trusting rank')
        out['weighted'].update(icc=None, excess=None)
        out['ranking_excess'] = None
        out['verdict'] = 'MALFORMED'
    elif all_ok:
        w_icc = sum(probs[k] / psum * out['scenarios'][k]['icc'] for k in names)
        w_exc = w_icc - w_ke
        out['weighted'].update(icc=round(w_icc, 4), excess=round(w_exc, 4))
        out['ranking_excess'] = round(w_exc, 4)
        out['verdict'] = _verdict(w_exc, out['prob_mix_feasible'])
    else:
        # at least one scenario has no ICC solution -> weighted ICC undefined;
        # the per-scenario statuses carry the signal (e.g. overvalued_past_floor).
        out['weighted'].update(icc=None, excess=None)
        out['ranking_excess'] = None
        bad = [k for k in names if out['scenarios'][k]['status'] != 'ok']
        out['flags'].append(
            f'no ICC solution for {", ".join(bad)} '
            f'({", ".join(out["scenarios"][k]["status"] for k in bad)})')
        # if every scenario is overvalued past the floor, that is itself a clean avoid
        if all(out['scenarios'][k]['status'] == 'overvalued_past_floor' for k in names):
            out['verdict'] = 'AVOID'
        else:
            out['verdict'] = 'CHECK'

    if not out['prob_mix_feasible'] and out['verdict'] not in ('MALFORMED',):
        out['flags'].append(
            f'price {price:.2f} exceeds the most optimistic scenario IVPS '
            f'{out["max_scenario_ivps"]:.2f} — outside the scenario hull')
    return out


def _verdict(weighted_excess, feasible):
    """Heuristic label off the ranking statistic. The number is what matters; the
    label is a reading aid. ~+/-1% band around the hurdle = roughly fair."""
    if not feasible:
        return 'AVOID'
    if weighted_excess >= 0.01:
        return 'LEAN BUY'
    if weighted_excess <= -0.01:
        return 'AVOID'
    return 'FAIR'


# --------------------------------------------------------------- price substrate

def _marks_path(ticker_dir):
    return os.path.join(ticker_dir, 'marks.jsonl')


def read_price(ticker_dir):
    """Latest sane mark as (price, date, source), or None if no marks file yet."""
    path = _marks_path(ticker_dir)
    if not os.path.exists(path):
        return None
    last = None
    with open(path) as f:
        for line in f:
            line = line.strip()
            if line:
                last = json.loads(line)
    if last is None:
        return None
    return last['price'], last['date'], last.get('source', '')


ICC_TITLE = 'IMPLIED RETURNS (ICC) — live (reactive to the price cell)'


def build_reactive_block(ws, ticker, scenarios, model_fmt='Model_{}',
                         price_ref='$B$5', prob_ref='Scenario_Summary!$B$5:$D$5',
                         names=('Bear', 'Base', 'Bull'), step=0.0025):
    """Write a LIVE, formula-driven ICC block into worksheet `ws` (no save here —
    the caller owns the workbook). Unlike a stamped value, this recomputes whenever
    the price cell `price_ref` (or any model assumption) changes — no macros.

    Method: ICC is the root of IVPS(r) = price, which has no closed form, but the
    operating forecast (NI_t, beginning equity_t) is independent of r. So for each
    scenario we tabulate IVPS(r) on a fine descending-rate grid (literal rates;
    IVPS by formula off the Model tab's NI row 40 and beginning-equity row 27), then
    INTERPOLATE the rate where the curve crosses the price. ~0.5bp vs the Python
    bisection at a 0.25% step.

    IVPS(r) = (BV0 + Σ_t (NI_t − r·Eq_{t-1})/(1+r)^t
                   + (NI_10 − r·Eq_10)(1+g)/(r−g)/(1+r)^10) / shares
    with NI_t = Model!L40:U40, Eq_{t-1} = Model!L27:U27, BV0 = Model!L48,
    shares = Model!L50, g = each scenario's terminal growth.

    Idempotent: re-running finds the existing block by its title and overwrites it.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.utils import get_column_letter

    BOLD = Font(bold=True)
    SECT = PatternFill('solid', start_color='F2F2F2')
    HDR = PatternFill('solid', start_color='D9E2F3')
    ITAL = Font(italic=True, size=9)
    WRAP = Alignment(wrap_text=True, vertical='top')
    GREENF = Font(color='1B7F3B')
    REDF = Font(color='C0392B')
    PCT, MONEY = '0.00%', '#,##0.00'

    # locate / clear a prior block, else append after current content
    start = None
    for rr in range(1, ws.max_row + 1):
        if str(ws.cell(row=rr, column=1).value or '').startswith('IMPLIED RETURNS (ICC)'):
            start = rr
            break
    if start is not None:
        for mr in list(ws.merged_cells.ranges):      # unmerge prior block's merges first
            if mr.min_row >= start:                   # (else MergedCell value is read-only)
                ws.unmerge_cells(str(mr))
        for rr in range(start, ws.max_row + 1):
            for cc in range(1, 16):
                ws.cell(row=rr, column=cc).value = None
    else:
        start = (ws.max_row + 2) if ws.max_row > 1 else 1

    # per-scenario rate grids (descending: r=0.30 .. g+0.004), equal length
    min_g = min(scenarios[n]['terminalGrowth'] for n in names)
    N = int(round((0.30 - (min_g + 0.004)) / step)) + 1
    grids = {}
    for n in names:
        g = scenarios[n]['terminalGrowth']
        lo = g + 0.004
        grids[n] = [round(0.30 - i * (0.30 - lo) / (N - 1), 6) for i in range(N)]

    # ---- helper grid (below the visible table): cols H/I, J/K, L/M
    helper_hdr = start + 14
    g0 = helper_hdr + 1
    g1 = g0 + N - 1
    helper_cols = {'Bear': ('H', 'I'), 'Base': ('J', 'K'), 'Bull': ('L', 'M')}
    ws.cell(row=start + 12, column=1,
            value='IVPS-by-discount-rate grid — helper for the interpolation above (do not edit)').font = ITAL
    for n in names:
        rc, ic = helper_cols[n]
        ws[f'{rc}{helper_hdr}'] = f'r ({n})'; ws[f'{rc}{helper_hdr}'].font = BOLD
        ws[f'{ic}{helper_hdr}'] = f'IVPS ({n})'; ws[f'{ic}{helper_hdr}'].font = BOLD
        M = model_fmt.format(n)
        g = scenarios[n]['terminalGrowth']
        for i, rval in enumerate(grids[n]):
            row = g0 + i
            ws[f'{rc}{row}'] = rval
            ws[f'{rc}{row}'].number_format = PCT
            R = f'${rc}{row}'
            ivps = (f"=({M}!$L$48"
                    f"+SUMPRODUCT(({M}!$L$40:$U$40-{R}*{M}!$L$27:$U$27)"
                    f"/(1+{R})^{{1,2,3,4,5,6,7,8,9,10}})"
                    f"+({M}!$U$40-{R}*{M}!$U$27)*(1+{g})/({R}-{g})/(1+{R})^10)"
                    f"/{M}!$L$50")
            ws[f'{ic}{row}'] = ivps
            ws[f'{ic}{row}'].number_format = MONEY

    # ---- visible table
    ws.cell(row=start, column=1, value=ICC_TITLE).font = Font(bold=True, size=12)
    ws.cell(row=start, column=1).fill = SECT
    ws.cell(row=start + 1, column=1, value=(
        f'ICC = the discount rate at which each scenario’s IVPS equals the price in '
        f'{price_ref.replace("$","")}. Recomputes when you change the price or any model '
        f'assumption (interpolated from the grid below; no macros). Excess = ICC − Ke: '
        f'positive ⇒ the price implies a return above our hurdle (cheap), negative ⇒ below it.')
    ).font = ITAL
    ws.merge_cells(start_row=start + 1, start_column=1, end_row=start + 1, end_column=7)
    ws.cell(row=start + 1, column=1).alignment = WRAP

    hdr = start + 3
    for j, h in enumerate(['Scenario', 'IVPS', 'Ke (hurdle)', 'ICC', 'ICC − Ke', 'Read']):
        c = ws.cell(row=hdr, column=1 + j, value=h); c.font = BOLD; c.fill = HDR
    row0 = hdr + 1
    for i, n in enumerate(names):
        r = row0 + i
        M = model_fmt.format(n)
        rc, ic = helper_cols[n]
        rng_r = f'${rc}${g0}:${rc}${g1}'
        rng_i = f'${ic}${g0}:${ic}${g1}'
        m = f'MATCH({price_ref},{rng_i},1)'
        interp = (f'INDEX({rng_r},{m})+({price_ref}-INDEX({rng_i},{m}))'
                  f'/(INDEX({rng_i},{m}+1)-INDEX({rng_i},{m}))'
                  f'*(INDEX({rng_r},{m}+1)-INDEX({rng_r},{m}))')
        ws.cell(row=r, column=1, value=n)
        ws.cell(row=r, column=2, value=f'={M}!$L$51').number_format = MONEY
        ws.cell(row=r, column=3, value=f'={M}!$B$5').number_format = PCT
        ws.cell(row=r, column=4, value=f'=IFERROR({interp},NA())').number_format = PCT
        ws.cell(row=r, column=5, value=f'=IF(ISNA(D{r}),"",D{r}-C{r})').number_format = PCT
        ws.cell(row=r, column=6, value=(
            f'=IF(ISNA(D{r}),IF({price_ref}>MAX({rng_i}),'
            f'"price above scenario — overvalued, no positive-excess solution",'
            f'"price below grid (ICC>30%)"),'
            f'IF(E{r}>0,"undervalued — price implies return above hurdle",'
            f'"overvalued — price implies return below hurdle"))')).alignment = WRAP

    wr = row0 + len(names)
    ws.cell(row=wr, column=1, value='Probability-weighted').font = BOLD
    ws.cell(row=wr, column=2, value=f'=SUMPRODUCT({prob_ref},B{row0}:B{wr-1})').number_format = MONEY
    ws.cell(row=wr, column=3, value=f'=SUMPRODUCT({prob_ref},C{row0}:C{wr-1})').number_format = PCT
    ws.cell(row=wr, column=4, value=f'=SUMPRODUCT({prob_ref},D{row0}:D{wr-1})').number_format = PCT
    ws.cell(row=wr, column=5, value=f'=D{wr}-C{wr}').number_format = PCT
    ws.cell(row=wr, column=6, value=(
        f'=IF(E{wr}>0.01,"LEAN BUY",IF(E{wr}<-0.01,"AVOID","FAIR"))'))
    for col in (1, 4, 5, 6):
        ws.cell(row=wr, column=col).font = BOLD

    note = wr + 2
    ws.cell(row=note, column=1,
            value='Ranking excess (weighted ICC − Ke) — the universe sort key:').font = BOLD
    ws.cell(row=note, column=4, value=f'=E{wr}').number_format = PCT
    ws.cell(row=note, column=4).font = BOLD
    ws.cell(row=note + 1, column=1, value='Price inside the scenario hull?')
    ws.cell(row=note + 1, column=4,
            value=f'=IF({price_ref}<=MAX(B{row0}:B{wr-1}),"yes","NO — above every scenario IVPS (avoid)")')

    # conditional formatting: green if excess > 0, red if < 0 (table + weighted)
    rng = f'E{row0}:E{wr}'
    ws.conditional_formatting.add(rng, CellIsRule(operator='greaterThan', formula=['0'], font=GREENF))
    ws.conditional_formatting.add(rng, CellIsRule(operator='lessThan', formula=['0'], font=REDF))

    for col, width in zip('ABCDEFG', (40, 12, 12, 10, 10, 52, 6)):
        ws.column_dimensions[col].width = width
    return wr


def append_mark(ticker_dir, date, price, source, allow_move=SANITY_MOVE):
    """Append a dated EOD mark. Append-only / never-backfill, mirroring
    consensus_snapshots.json: a same-date re-run REPLACES that date's row; a new
    date is appended. Sanity clamp: reject price<=0 or a >allow_move jump vs the
    last good mark (catches split/null/zero garbage from a flaky free feed).

    Returns (written: bool, reason: str). A rejected mark leaves the prior good
    mark in place — the feed degrades by going stale (visible via the date), never
    by silently overwriting a good price with a bad one.
    """
    if not (isinstance(price, (int, float)) and price > 0):
        return False, f'rejected: non-positive price {price!r}'
    path = _marks_path(ticker_dir)
    rows = []
    if os.path.exists(path):
        with open(path) as f:
            rows = [json.loads(l) for l in f if l.strip()]
    prior = [r for r in rows if r['date'] != date]
    last_good = prior[-1]['price'] if prior else None
    if last_good and abs(price / last_good - 1) > allow_move:
        return False, (f'rejected: {price:.2f} moves '
                       f'{(price/last_good-1)*100:+.0f}% vs last good {last_good:.2f}')
    prior.append({'date': date, 'price': round(float(price), 4), 'source': source})
    prior.sort(key=lambda r: r['date'])
    with open(path, 'w') as f:
        for r in prior:
            f.write(json.dumps(r) + '\n')
    return True, 'written'
