#!/usr/bin/env python3
"""Build/refresh the 'Valuation Multiples' workbook tab — the evolution of
P/E, P/forward-E, PEG, P/S and P/B across the fiscal-year window.

Methodology (documented on the tab):
  - Price = closing price 30 calendar days after that fiscal year's EARNINGS
    RELEASE (first 8-K filed after FYE, via EDGAR submissions), next trading
    day if markets closed. Post-digestion price, not announcement noise.
  - Prices from Yahoo are split-adjusted (current share basis), so as-filed
    EPS is converted to current basis via the cumulative split factor after
    the release date; diluted shares are derived as NI ÷ EPS (basis-proof).
  - P/E = price / EPS (trailing, as-filed FY). Forward P/E = price / NEXT
    fiscal year's REALIZED EPS (perfect-foresight forward); the newest year
    uses a consensus estimate if provided (blue cell, editable).
  - PEG = trailing P/E ÷ (100 × forward EPS growth); 'n/m' when growth ≤ 0
    or either EPS ≤ 0 — which is the point for cyclicals.
  - P/S = market cap / revenue;  P/B = market cap / common equity
    (market cap = price × derived diluted shares, current basis).

Facts-clause artifact: as-filed + as-traded data, appended automatically when
a new fiscal year lands. Usage:  python3 build_multiples.py [--consensus-eps X]
"""
import argparse, datetime, json, os, sys, urllib.request

HERE = os.path.dirname(os.path.abspath(__file__))
TICKER_DIR = os.path.dirname(HERE)
TICKER = os.path.basename(TICKER_DIR)
WB_PATH = os.path.join(TICKER_DIR, f'{TICKER}_Integrated_Financials.xlsx')
UA = os.environ.get('BAV_SEC_IDENTITY', 'BAV multiples contact@example.com')
PRICE_LAG_DAYS = 30


def edgar_8k_dates(cik):
    """All 8-K filing dates, newest first (submissions feed + paginated archives)."""
    dates = []
    def harvest(d):
        r = d if 'form' in d else d['filings']['recent']
        items = r.get('items', [''] * len(r['form']))
        dates.extend(dd for f, dd, it in zip(r['form'], r['filingDate'], items)
                     if f == '8-K' and '2.02' in (it or ''))
    url = f'https://data.sec.gov/submissions/CIK{int(cik):010d}.json'
    req = urllib.request.Request(url, headers={'User-Agent': UA})
    data = json.load(urllib.request.urlopen(req, timeout=30))
    harvest(data)
    for extra in data['filings'].get('files', []):
        u2 = 'https://data.sec.gov/submissions/' + extra['name']
        req2 = urllib.request.Request(u2, headers={'User-Agent': UA})
        harvest(json.load(urllib.request.urlopen(req2, timeout=30)))
    return sorted(set(dates))


def release_after(fye, dates8k):
    """First 8-K within 90 days after fiscal year end = the earnings release."""
    fye_s = fye.isoformat()
    for d in dates8k:
        if d > fye_s and (datetime.date.fromisoformat(d) - fye).days <= 90:
            return datetime.date.fromisoformat(d)
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--consensus-eps', type=float, default=None,
                    help='next-FY consensus diluted EPS (current basis) for the newest year')
    args = ap.parse_args()

    import openpyxl, yfinance as yf
    from openpyxl.styles import Font, PatternFill

    rowmap = json.load(open(os.path.join(HERE, 'rowmap.json')))
    wb = openpyxl.load_workbook(WB_PATH)
    is_ws = wb['Income Statement']
    bs_ws = wb['Balance Sheet']

    # locate rows by label (never assume positions)
    def find_row(ws, *needles, avoid=()):
        for r in range(1, ws.max_row + 1):
            lab = str(ws.cell(row=r, column=1).value or '').lower()
            if lab and all(n in lab for n in needles) and not any(a in lab for a in avoid):
                return r
        return None

    r_rev = find_row(is_ws, 'revenue') or find_row(is_ws, 'net sales') or find_row(is_ws, 'sales')
    r_ni = find_row(is_ws, 'net income', avoid=('noncontrolling', 'attributable to noncontrolling', 'per share'))
    r_eps = find_row(is_ws, 'diluted', 'per share') or find_row(is_ws, 'diluted earnings')
    r_eq = find_row(bs_ws, "total", "equity", avoid=('liabilities', 'including', 'noncontrolling'))
    assert all((r_rev, r_ni, r_eps, r_eq)), f'row discovery failed: rev={r_rev} ni={r_ni} eps={r_eps} eq={r_eq}'

    # year columns from the date header row (row 6)
    cols = []
    for c in range(2, 30):
        v = is_ws.cell(row=6, column=c).value
        if isinstance(v, (datetime.date, datetime.datetime)):
            cols.append((c, v.date() if isinstance(v, datetime.datetime) else v))
    assert cols, 'no date headers found'

    cik = {'GOOGL': 1652044, 'MU': 723125, 'MSFT': 789019, 'NVDA': 1045810}.get(TICKER)
    if cik is None:
        fm = open(os.path.join(TICKER_DIR, 'dossier.md')).read()
        import re
        cik = int(re.search(r'^cik:\s*(\d+)', fm, re.M).group(1))
    dates8k = edgar_8k_dates(cik)

    tk = yf.Ticker(TICKER)
    hist = tk.history(start=(cols[0][1] - datetime.timedelta(days=5)).isoformat(),
                      auto_adjust=False)['Close']
    hist.index = [d.date() for d in hist.index]
    splits = tk.splits  # Series indexed by date, value = split ratio

    def split_factor_after(d):
        f = 1.0
        for sd, ratio in splits.items():
            if sd.date() > d and float(ratio) > 0:
                f *= float(ratio)
        return f

    def price_on_or_after(d):
        for k in range(8):
            dd = d + datetime.timedelta(days=k)
            if dd in hist.index:
                return float(hist.loc[dd]), dd
        return None, None

    import re as _re

    def numval(ws, r, c, depth=0):
        """Numeric cell value; evaluates simple subtotal formulas (=SUM(range), =A7+A8-A9)."""
        v = ws.cell(row=r, column=c).value
        if isinstance(v, (int, float)):
            return v
        if not (isinstance(v, str) and v.startswith('=')) or depth > 3:
            return None
        m = _re.fullmatch(r'=SUM\((\$?[A-Z]+\$?\d+):(\$?[A-Z]+\$?\d+)\)', v.replace(' ', ''))
        if m:
            a, b = (_re.match(r'\$?([A-Z]+)\$?(\d+)', x).groups() for x in m.groups())
            total = 0
            for rr in range(int(a[1]), int(b[1]) + 1):
                vv = numval(ws, rr, c, depth + 1)
                total += vv or 0
            return total
        toks = _re.findall(r'([+\-]?)\$?([A-Z]+)\$?(\d+)', v[1:].replace(' ', ''))
        if toks and _re.fullmatch(r'([+\-]?\$?[A-Z]+\$?\d+)+', v[1:].replace(' ', '')):
            total = 0
            for sign, colL, rr in toks:
                vv = numval(ws, int(rr), openpyxl.utils.column_index_from_string(colL), depth + 1)
                total += (-1 if sign == '-' else 1) * (vv or 0)
            return total
        return None

    years = []
    for c, fye in cols:
        cell = lambda ws, r: numval(ws, r, c)
        rel = release_after(fye, dates8k)
        price, pdate = price_on_or_after(rel + datetime.timedelta(days=PRICE_LAG_DAYS)) if rel else (None, None)
        eps_raw, ni, rev, eq = cell(is_ws, r_eps), cell(is_ws, r_ni), cell(is_ws, r_rev), cell(bs_ws, r_eq)
        fac = split_factor_after(rel) if rel else 1.0
        eps = eps_raw if isinstance(eps_raw, (int, float)) and eps_raw else None
        shares = (ni / eps) if (eps and ni) else None
        years_basis_fix = True   # resolved after the loop, using newest-year shares
        years.append(dict(fy=fye, col=c, release=rel, pdate=pdate, price=price,
                          eps=eps, ni=ni, rev=rev, eq=eq, shares=shares))

    # EPS basis fix: some workbook years carry as-filed (pre-split) EPS, others
    # already split-adjusted. For each year choose raw vs raw/split-factor by which
    # implied share count is closer (log-scale) to the newest year's implied shares.
    import math
    ref = next((y['ni'] / y['eps'] for y in reversed(years) if y.get('eps') and y.get('ni')), None)
    if ref:
        for y in years:
            if not (y.get('eps') and y.get('ni')):
                continue
            fac = split_factor_after(y['release']) if y['release'] else 1.0
            cands = [(y['eps'], y['ni'] / y['eps'])]
            if fac != 1.0:
                cands.append((y['eps'] / fac, y['ni'] / (y['eps'] / fac)))
            eps_best, sh_best = min(cands, key=lambda t: abs(math.log(t[1] / ref)))
            y['eps'], y['shares'] = eps_best, sh_best

    # forward EPS = next year's realized; newest year = consensus arg (estimate)
    for i, y in enumerate(years):
        y['fwd_eps'] = years[i + 1]['eps'] if i + 1 < len(years) else args.consensus_eps
        y['fwd_is_est'] = i + 1 >= len(years)

    # ---- write the tab
    if 'Valuation Multiples' in wb.sheetnames:
        del wb['Valuation Multiples']
    ws = wb.create_sheet('Valuation Multiples')
    BOLD, BLUE, ITAL = Font(bold=True), Font(color='0000FF'), Font(italic=True, size=9)
    YEL = PatternFill('solid', start_color='FFF2CC')
    ws['A1'] = f'{TICKER} — Valuation Multiples by Fiscal Year'
    ws['A1'].font = Font(bold=True, size=13)
    ws['A2'] = (f'Price = close {PRICE_LAG_DAYS} calendar days after the FY earnings release '
                '(first post-FYE 8-K). Forward E = next FY realized EPS (newest year: consensus, blue). '
                'PEG = P/E ÷ fwd EPS growth%; n/m when growth or EPS ≤ 0. As-filed fundamentals; '
                'split-consistent basis. Facts tab — auto-extends each new fiscal year.')
    ws['A2'].font = ITAL
    rows = ['Fiscal year end', 'Earnings release (8-K)', f'Price date (+{PRICE_LAG_DAYS}d)',
            'Price ($, split-adj basis)', 'Diluted EPS ($)', 'P/E (trailing)',
            'Forward EPS ($)', 'P/forward E', 'Fwd EPS growth %', 'PEG',
            'Market cap ($M)', 'Revenue ($M)', 'P/S', 'Common equity ($M)', 'P/B']
    for i, lab in enumerate(rows):
        ws.cell(row=4 + i, column=1, value=lab).font = BOLD if lab.startswith(('P/', 'PEG')) else Font()
    ws.column_dimensions['A'].width = 30
    for j, y in enumerate(years):
        col = 2 + j
        L = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[L].width = 13
        def put(r, v, fmt='#,##0.00', est=False):
            cell = ws.cell(row=r, column=col, value=v)
            cell.number_format = fmt
            if est:
                cell.font = BLUE
            return cell
        put(4, y['fy'], 'MMM DD, YYYY').font = BOLD
        put(5, y['release'], 'MMM DD, YYYY')
        put(6, y['pdate'], 'MMM DD, YYYY')
        put(7, y['price'])
        put(8, round(y['eps'], 2) if y['eps'] else None)
        pe = (y['price'] / y['eps']) if (y['price'] and y['eps'] and y['eps'] > 0) else None
        put(9, round(pe, 1) if pe else 'n/m', '0.0').fill = YEL
        put(10, round(y['fwd_eps'], 2) if y['fwd_eps'] else None, est=y['fwd_is_est'])
        fpe = (y['price'] / y['fwd_eps']) if (y['price'] and y['fwd_eps'] and y['fwd_eps'] > 0) else None
        put(11, round(fpe, 1) if fpe else 'n/m', '0.0').fill = YEL
        g = ((y['fwd_eps'] / y['eps'] - 1) * 100) if (y['eps'] and y['fwd_eps'] and y['eps'] > 0) else None
        put(12, round(g, 1) if g is not None else 'n/m', '0.0')
        peg = (pe / g) if (pe and g and g > 0) else None
        put(13, round(peg, 2) if peg else 'n/m', '0.00').fill = YEL
        mcap = y['price'] * y['shares'] if (y['price'] and y['shares']) else None
        put(14, round(mcap) if mcap else None, '#,##0')
        put(15, y['rev'], '#,##0')
        put(16, round(mcap / y['rev'], 2) if (mcap and y['rev']) else None, '0.00').fill = YEL
        put(17, y['eq'], '#,##0')
        put(18, round(mcap / y['eq'], 2) if (mcap and y['eq']) else None, '0.00').fill = YEL
    wb.save(WB_PATH)
    print(f'{TICKER}: Valuation Multiples tab written, {len(years)} fiscal years')
    for y in years:
        pe = y['price'] / y['eps'] if (y['price'] and y['eps'] and y['eps'] > 0) else None
        print(f"  {y['fy']}  rel={y['release']}  px={y['price'] and round(y['price'],2)}  "
              f"eps={y['eps'] and round(y['eps'],2)}  P/E={pe and round(pe,1)}")


if __name__ == '__main__':
    sys.exit(main())
