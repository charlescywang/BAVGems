#!/usr/bin/env python3
"""
NVDA Integrated Financials Workbook Builder — v3
Follows updated BAV Pipeline SKILL.md with:
  - SUMIF-based classification with dropdown menus
  - ⚠️ flags on ambiguous items
  - CoE as formula (=Rf+Beta*ERP)
  - Formula-linked tabs throughout
  - Proper formatting standards
"""

import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── Color constants (xlsx skill color coding) ──
BLUE_FONT = Font(name='Arial', color='0000FF')          # hardcoded inputs
BLACK_FONT = Font(name='Arial', color='000000')          # formulas
GREEN_FONT = Font(name='Arial', color='008000')          # cross-sheet links
GRAY_FONT = Font(name='Arial', color='999999')           # computed / derived
BOLD_FONT = Font(name='Arial', bold=True)
BOLD_BLUE = Font(name='Arial', bold=True, color='0000FF')
HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')
SECTION_FILL = PatternFill('solid', fgColor='E2EFDA')
CHECK_FILL = PatternFill('solid', fgColor='FFF2CC')
NUM_FMT = '#,##0;(#,##0);"-"'
PCT_FMT = '0.0%'
PRICE_FMT = '$#,##0.00'
THIN_BORDER = Border(bottom=Side(style='thin'))

from datetime import date
# Fiscal year-end dates (from NVIDIA 10-K filings)
FY_END_DATES = [date(2022, 1, 30), date(2023, 1, 29), date(2024, 1, 28), date(2025, 1, 26), date(2026, 1, 25)]
YEARS = ['FY2022', 'FY2023', 'FY2024', 'FY2025', 'FY2026']  # kept as string keys for dicts
DATA_COLS = ['B', 'C', 'D', 'E', 'F']  # IS/BS/CF data columns
CF_DATA_COLS = ['C', 'D', 'E', 'F', 'G']  # Condensed Financials data columns (offset by 1)
DATE_FMT = 'MMM DD, YYYY'  # Excel date format for headers

def set_col_widths(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def write_header_block(ws, company, statement, row=1):
    ws.cell(row=row, column=1, value=f'Company: {company}').font = BOLD_FONT
    ws.cell(row=row+1, column=1, value=f'Statement: {statement}').font = Font(name='Arial', italic=True)
    ws.cell(row=row+2, column=1, value='Units: USD in Millions').font = Font(name='Arial', italic=True)
    ws.cell(row=row+3, column=1, value='Source: SEC 10-K Filings, Press Releases').font = Font(name='Arial', italic=True)

def write_year_headers(ws, row, cols, years, dates=None):
    """Write year headers. If dates provided, use date objects for proper YEAR() support."""
    ws.cell(row=row, column=1, value='Line Item').font = BOLD_FONT
    for i, col_letter in enumerate(cols):
        c = ws.cell(row=row, column=ord(col_letter)-64)
        if dates:
            c.value = dates[i]
            c.number_format = DATE_FMT
        else:
            c.value = years[i]
        c.font = BOLD_FONT
        c.alignment = Alignment(horizontal='center')
        c.fill = HEADER_FILL

def write_data_row(ws, row, label, values, cols, indent=False, font=None, num_fmt=NUM_FMT):
    lbl = f'  {label}' if indent else label
    ws.cell(row=row, column=1, value=lbl).font = font or Font(name='Arial')
    for col_letter, val in zip(cols, values):
        c = ws.cell(row=row, column=ord(col_letter)-64, value=val)
        c.font = font or BLUE_FONT
        c.number_format = num_fmt
        c.alignment = Alignment(horizontal='right')

# ════════════════════════════════════════════════════════════════════
# TAB 1: INCOME STATEMENT
# ════════════════════════════════════════════════════════════════════
ws_is = wb.active
ws_is.title = 'Income Statement'
set_col_widths(ws_is, {'A': 48, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16})

write_header_block(ws_is, 'NVIDIA Corporation (NVDA)', 'Consolidated Statements of Income')
write_year_headers(ws_is, 6, DATA_COLS, YEARS, dates=FY_END_DATES)

is_data = [
    (7,  'Revenue',                        [26914, 26974, 60922, 130497, 215938]),
    (8,  'Cost of Revenue',                [-9439, -11618, -16621, -29168, -62475]),
    (9,  'Gross Profit',                   [17475, 15356, 44301, 101329, 153463]),
    (11, 'Research & Development',         [-5268, -7339, -8675, -12893, -18316]),
    (12, 'Sales, General & Administrative',[-2166, -2440, -2654, -3481, -4760]),
    (13, 'Total Operating Expenses',       [-7434, -9779, -11329, -16374, -23076]),
    (14, 'Operating Income',               [10041, 5577, 32972, 84955, 130387]),
    (16, 'Interest Income',                [42, 267, 866, 2539, 3800]),
    (17, 'Interest Expense',               [-236, -257, -257, -246, -246]),
    (18, 'Other Income (Expense), Net',    [107, -43, 237, 421, 500]),
    (19, 'Income Before Income Taxes',     [9954, 5544, 33818, 87669, 134441]),
    (20, 'Income Tax Expense',             [-189, -1861, -4042, -14527, -14374]),
    (21, 'Net Income',                     [9752, 4368, 29760, 72880, 120067]),
    (23, 'EPS - Basic',                    [3.91, 1.76, 12.06, 2.97, 4.94]),
    (24, 'EPS - Diluted',                  [3.85, 1.74, 11.93, 2.94, 4.90]),
    (25, 'Basic Shares Outstanding (M)',   [2496, 2487, 24660, 24530, 24300]),
    (26, 'Diluted Shares Outstanding (M)', [2535, 2507, 24940, 24490, 24500]),
]
for row, label, vals in is_data:
    fmt = '#,##0.00' if row in (23, 24) else NUM_FMT
    write_data_row(ws_is, row, label, vals, DATA_COLS, num_fmt=fmt)

# Add note about FY2026 being estimated
ws_is.cell(row=28, column=1, value="Note: FY2026 data is estimated.").font = Font(name='Arial', italic=True, color='999999')

# Bold subtotals
for r in [9, 13, 14, 19, 21]:
    ws_is.cell(row=r, column=1).font = BOLD_FONT

# ════════════════════════════════════════════════════════════════════
# TAB 2: BALANCE SHEET
# ════════════════════════════════════════════════════════════════════
ws_bs = wb.create_sheet('Balance Sheet')
set_col_widths(ws_bs, {'A': 48, 'B': 16, 'C': 16, 'D': 16, 'E': 16})

write_header_block(ws_bs, 'NVIDIA Corporation (NVDA)', 'Consolidated Balance Sheets')
write_year_headers(ws_bs, 6, DATA_COLS, YEARS, dates=FY_END_DATES)

bs_data = [
    (7,  'ASSETS',                                    None),
    (8,  'Cash and Cash Equivalents',                 [1990, 3389, 7280, 8495, 8500]),
    (9,  'Short-term Investments',                    [19218, 9907, 18704, 22417, 54100]),
    (10, 'Accounts Receivable, Net',                  [4650, 3827, 9999, 17484, 29000]),
    (11, 'Inventories',                               [2605, 5159, 5282, 8076, 22000]),
    (12, 'Prepaid Expenses & Other Current Assets',   [791, 1389, 2159, 3482, 5800]),
    (13, 'Total Current Assets',                      [28254, 23671, 43424, 59954, 119400]),
    (15, 'Property, Plant and Equipment, Net',        [3213, 3807, 4885, 6507, 11200]),
    (16, 'Operating Lease Right-of-Use Assets',       [829, 1038, 1346, 1795, 2800]),
    (17, 'Goodwill',                                  [4372, 4372, 4430, 4453, 4453]),
    (18, 'Intangible Assets, Net',                    [2339, 1676, 1112, 765, 500]),
    (19, 'Deferred Income Tax Assets',                [4019, 3396, 6081, 9578, 18000]),
    (20, 'Other Long-term Assets',                    [2928, 3820, 5378, 7509, 22000]),
    (21, 'Total Assets',                              [44187, 41182, 65728, 96013, 178353]),
    (23, 'LIABILITIES',                               None),
    (24, 'Accounts Payable',                          [1783, 1193, 2699, 5353, 12000]),
    (25, 'Accrued and Other Current Liabilities',     [4686, 5547, 7567, 10613, 22000]),
    (26, 'Short-term Debt',                           [0, 1250, 1250, 1249, 1249]),
    (27, 'Deferred Revenue (Current)',                [488, 354, 764, 1426, 3000]),
    (28, 'Total Current Liabilities',                 [7496, 8885, 12909, 19482, 38249]),
    (30, 'Long-term Debt',                            [10946, 9703, 8459, 8462, 8462]),
    (31, 'Long-term Operating Lease Liabilities',     [741, 902, 1119, 1490, 2400]),
    (32, 'Other Long-term Liabilities',               [1517, 1655, 2217, 3222, 5000]),
    (33, 'Total Liabilities',                         [20700, 21145, 24704, 32656, 54111]),
    (35, 'STOCKHOLDERS EQUITY',                       None),
    (36, 'Common Stock and Additional Paid-in Capital',[10385, 11971, 13132, 21168, 29000]),
    (37, 'Retained Earnings',                         [16235, 10171, 29817, 43927, 96000]),
    (38, 'Accumulated Other Comprehensive Income (Loss)',[-133, -106, -109, -262, -758]),
    (39, 'Total Stockholders Equity',                 [23487, 20037, 41024, 63033, 124242]),
    (41, 'Total Liabilities and Stockholders Equity', [44187, 41182, 65728, 96013, 178353]),
]
for row, label, vals in bs_data:
    if vals is None:
        ws_bs.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws_bs.cell(row=row, column=1).fill = SECTION_FILL
    else:
        write_data_row(ws_bs, row, label, vals, DATA_COLS)
for r in [13, 21, 28, 33, 39, 41]:
    ws_bs.cell(row=r, column=1).font = BOLD_FONT

# ════════════════════════════════════════════════════════════════════
# TAB 3: CASH FLOW
# ════════════════════════════════════════════════════════════════════
ws_cf = wb.create_sheet('Cash Flow')
set_col_widths(ws_cf, {'A': 48, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16})

write_header_block(ws_cf, 'NVIDIA Corporation (NVDA)', 'Consolidated Statements of Cash Flows')
write_year_headers(ws_cf, 6, DATA_COLS, YEARS, dates=FY_END_DATES)

cf_data = [
    (7,  'OPERATING ACTIVITIES',              None),
    (8,  'Net Income',                        [9752, 4368, 29760, 72880, 120067]),
    (9,  'Depreciation and Amortization',     [1174, 1544, 1957, 2350, 3200]),
    (10, 'Stock-based Compensation',          [2004, 2709, 3746, 5008, 7000]),
    (11, 'Other Non-Cash Adjustments',        [-3000, 1340, -1700, -5000, -15549]),
    (12, 'Changes in Working Capital',        [-701, -3289, -4343, -8500, -12000]),
    (13, 'Net Cash from Operating Activities',[9108, 5641, 28943, 64249, 102718]),
    (15, 'INVESTING ACTIVITIES',              None),
    (16, 'Capital Expenditures',              [-976, -1833, -1069, -3234, -5718]),
    (17, 'Purchases of Investments',          [-24189, -11897, -24869, -31055, -82000]),
    (18, 'Proceeds from Sales of Investments',[20042, 19953, 15539, 20796, 30000]),
    (19, 'Other Investing Activities',        [-1225, -110, -367, -987, -2000]),
    (20, 'Net Cash from Investing Activities',[-6348, 6113, -10766, -14480, -59718]),
    (22, 'FINANCING ACTIVITIES',              None),
    (23, 'Repurchases of Common Stock',       [-5762, -10039, -9533, -32666, -40400]),
    (24, 'Dividends Paid',                    [-399, -398, -395, -589, -974]),
    (25, 'Debt Repayments',                   [0, -1, -1250, 0, 0]),
    (26, 'Other Financing Activities',        [-2577, -1671, 3456, -3735, -1621]),
    (27, 'Net Cash from Financing Activities',[-8738, -12109, -7722, -36990, -42995]),
    (29, 'Net Change in Cash',                [-5887, -381, 10405, 12807, 5]),
    (30, 'Cash at Beginning of Period',       [7877, 1990, 3389, 7280, 8495]),
    (31, 'Cash at End of Period',             [1990, 3389, 7280, 8495, 8500]),
]
for row, label, vals in cf_data:
    if vals is None:
        ws_cf.cell(row=row, column=1, value=label).font = BOLD_FONT
        ws_cf.cell(row=row, column=1).fill = SECTION_FILL
    else:
        write_data_row(ws_cf, row, label, vals, DATA_COLS)
for r in [13, 20, 27, 29]:
    ws_cf.cell(row=r, column=1).font = BOLD_FONT

# ════════════════════════════════════════════════════════════════════
# TAB 4: CONDENSED FINANCIALS (with SUMIF + dropdowns)
# ════════════════════════════════════════════════════════════════════
ws_cond = wb.create_sheet('Condensed Financials')
set_col_widths(ws_cond, {'A': 48, 'B': 28, 'C': 16, 'D': 16, 'E': 16, 'F': 16})

# Header
ws_cond['A1'] = 'NVIDIA Corporation — Condensed Financials'
ws_cond['A1'].font = BOLD_FONT
ws_cond['A2'] = 'All values linked to source tabs via formulas'
ws_cond['A2'].font = Font(name='Arial', italic=True)
ws_cond['A3'] = 'Classification: User-approved (ROU=Operating Asset, Lease Liab=Financial)'
ws_cond['A3'].font = Font(name='Arial', italic=True, color='666666')

# Year headers (Col C-F for data, Col B for classification/spacer)
# Use date objects so YEAR() works downstream
for i, dt in enumerate(FY_END_DATES):
    col = i + 3  # C=3, D=4, E=5, F=6
    c = ws_cond.cell(row=5, column=col, value=dt)
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal='center')
    c.fill = HEADER_FILL
    c.number_format = DATE_FMT

# ── SECTION A: CONDENSED INCOME STATEMENT ──
ws_cond['A7'] = 'A. CONDENSED INCOME STATEMENT'
ws_cond['A7'].font = BOLD_FONT
ws_cond['A7'].fill = SECTION_FILL

# Row map for IS references (IS data in cols B-E → 'Income Statement'!{col}{row})
IS_COLS = {'FY2022': 'B', 'FY2023': 'C', 'FY2024': 'D', 'FY2025': 'E', 'FY2026': 'F'}
BS_COLS = IS_COLS  # same

R_NI = 9
R_INT_EXP = 10
R_INT_INC = 11
R_NET_INT = 12
R_PRETAX = 13
R_TAX = 14
R_ETR = 15
R_NIAT = 16
R_NOPAT = 17

cf_labels = {
    R_NI: 'Net Income',
    R_INT_EXP: 'Interest Expense (abs)',
    R_INT_INC: 'Interest Income',
    R_NET_INT: 'Net Interest Expense',
    R_PRETAX: 'Pretax Income',
    R_TAX: 'Tax Expense (abs)',
    R_ETR: 'Effective Tax Rate',
    R_NIAT: 'Net Interest After Tax',
    R_NOPAT: 'NOPAT',
}
for r, lbl in cf_labels.items():
    ws_cond.cell(row=r, column=1, value=lbl).font = Font(name='Arial')

# IS row references
IS_NI = 21; IS_INT_INC = 16; IS_INT_EXP = 17; IS_PRETAX = 19; IS_TAX = 20

for i, yr in enumerate(YEARS):
    dc = IS_COLS[yr]  # data col in IS tab
    cc = get_column_letter(i + 3)  # col in this tab (C, D, E, F)

    # Net Income
    f = f"='Income Statement'!{dc}{IS_NI}"
    c = ws_cond.cell(row=R_NI, column=i+3); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Interest Expense (abs)
    f = f"=ABS('Income Statement'!{dc}{IS_INT_EXP})"
    c = ws_cond.cell(row=R_INT_EXP, column=i+3); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Interest Income
    f = f"='Income Statement'!{dc}{IS_INT_INC}"
    c = ws_cond.cell(row=R_INT_INC, column=i+3); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Net Interest Expense
    f = f"={cc}{R_INT_EXP}-{cc}{R_INT_INC}"
    c = ws_cond.cell(row=R_NET_INT, column=i+3); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Pretax Income
    f = f"='Income Statement'!{dc}{IS_PRETAX}"
    c = ws_cond.cell(row=R_PRETAX, column=i+3); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Tax Expense (abs)
    f = f"=ABS('Income Statement'!{dc}{IS_TAX})"
    c = ws_cond.cell(row=R_TAX, column=i+3); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # ETR
    f = f"=IFERROR({cc}{R_TAX}/{cc}{R_PRETAX},0)"
    c = ws_cond.cell(row=R_ETR, column=i+3); c.value = f; c.font = BLACK_FONT; c.number_format = PCT_FMT

    # Net Interest After Tax
    f = f"={cc}{R_NET_INT}*(1-{cc}{R_ETR})"
    c = ws_cond.cell(row=R_NIAT, column=i+3); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # NOPAT
    f = f"={cc}{R_NI}+{cc}{R_NIAT}"
    c = ws_cond.cell(row=R_NOPAT, column=i+3); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

ws_cond.cell(row=R_NOPAT, column=1).font = BOLD_FONT

# ── SECTION C: CLASSIFICATION TABLE (built first so SUMIF can reference it) ──
# Classification table: Col A=Label, Col B=Classification(dropdown), Col C-F=FY values
CLASS_START = 63
ws_cond['A60'] = 'C. BALANCE SHEET CLASSIFICATION TABLE'
ws_cond['A60'].font = BOLD_FONT
ws_cond['A60'].fill = SECTION_FILL

ws_cond.cell(row=62, column=1, value='Line Item').font = BOLD_FONT
ws_cond.cell(row=62, column=2, value='Classification').font = BOLD_FONT
for i, dt in enumerate(FY_END_DATES):
    c = ws_cond.cell(row=62, column=i+3, value=dt)
    c.font = BOLD_FONT
    c.alignment = Alignment(horizontal='center')
    c.fill = HEADER_FILL
    c.number_format = DATE_FMT

# BS line items with their default classifications and BS row references
# Format: (label, default_classification, BS_row, is_ambiguous, flag_text)
bs_items = [
    ('Cash and cash equivalents',          'Financial Asset',            8,  False, ''),
    ('Short-term investments',             'Financial Asset',            9,  False, ''),
    ('Accounts receivable',                'Op. Working Capital Asset',  10, False, ''),
    ('Inventories',                        'Op. Working Capital Asset',  11, False, ''),
    ('Prepaid expenses & other current',   'Op. Working Capital Asset',  12, False, ''),
    ('Property, plant & equipment',        'Op. Long-Term Asset',        15, False, ''),
    ('Operating lease ROU assets',         'Op. Long-Term Asset',        16, True,  '⚠️ (User: Operating)'),
    ('Goodwill',                           'Op. Long-Term Asset',        17, False, ''),
    ('Intangible assets',                  'Op. Long-Term Asset',        18, False, ''),
    ('Deferred income tax assets',         'Op. Long-Term Asset',        19, False, ''),
    ('Other long-term assets',             'Op. Long-Term Asset',        20, False, ''),
    ('Accounts payable',                   'Op. Working Capital Liability', 24, False, ''),
    ('Accrued & current liabilities',      'Op. Working Capital Liability', 25, False, ''),
    ('Deferred revenue (current)',         'Op. Working Capital Liability', 27, False, ''),
    ('Short-term debt',                    'Financial Liability',        26, False, ''),
    ('Long-term debt',                     'Financial Liability',        30, False, ''),
    ('Operating lease liabilities (LT)',   'Financial Liability',        31, True,  '⚠️ (User: Financial)'),
    ('Other long-term liabilities',        'Op. Long-Term Liability',    32, False, ''),
]

CLASS_END = CLASS_START + len(bs_items) - 1  # 63 + 17 = 80

for idx, (label, classification, bs_row, is_ambiguous, flag) in enumerate(bs_items):
    r = CLASS_START + idx

    # Col A: Label
    ws_cond.cell(row=r, column=1, value=label).font = Font(name='Arial')

    # Col B: Classification (will have dropdown)
    class_val = classification
    if is_ambiguous:
        class_val = f"{classification} {flag}"
    ws_cond.cell(row=r, column=2, value=class_val).font = BLUE_FONT

    # Col C-F: FY values (formulas to Balance Sheet)
    for i, yr in enumerate(YEARS):
        dc = BS_COLS[yr]
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=r, column=i+3)
        c.value = f
        c.font = GREEN_FONT
        c.number_format = NUM_FMT

# Add Data Validation dropdown to classification cells
# Note: The dropdown has the 6 clean values. Items with ⚠️ flags have extended text
# but the dropdown still works — user can select a clean value to replace the flagged one.
categories = '"Op. Working Capital Asset,Op. Working Capital Liability,Op. Long-Term Asset,Op. Long-Term Liability,Financial Asset,Financial Liability"'
dv = DataValidation(type="list", formula1=categories, allow_blank=False)
dv.error = "Please select a valid classification"
dv.errorTitle = "Invalid Classification"
dv.prompt = "Select the balance sheet classification for this item"
dv.promptTitle = "Classification"
dv.showErrorMessage = True
dv.showInputMessage = True
for r in range(CLASS_START, CLASS_END + 1):
    dv.add(ws_cond.cell(row=r, column=2))
ws_cond.add_data_validation(dv)

# ── SECTION B: CONDENSED BALANCE SHEET (with SUMIF from classification table) ──
ws_cond['A19'] = 'B. CONDENSED BALANCE SHEET'
ws_cond['A19'].font = BOLD_FONT
ws_cond['A19'].fill = SECTION_FILL

# For SUMIF: criteria range is $B$63:$B$80, sum range is {col}$63:{col}$80
# But ⚠️ flagged items have extended text, so we need to handle matching.
# Solution: Use SUMIF with wildcard matching for categories that might have flags.
# Actually, simpler: use SUMIFS or just match the exact string.
# Since flagged items contain the base category + " ⚠️ ...", we use SUMIF with "*" wildcard:
# =SUMIF($B$63:$B$80,"Op. Long-Term Asset*",C63:C80) matches both
# "Op. Long-Term Asset" and "Op. Long-Term Asset ⚠️ (User: Operating)"

# But SUMIF wildcards: "*" matches any sequence. So "Op. Long-Term Asset*" would match.
# Actually in Excel, the pattern needs to be exact or use wildcards at the end.
# "Op. Working Capital Asset*" would match "Op. Working Capital Asset" and
# "Op. Working Capital Asset ⚠️ ..." — perfect.

# Row layout for Condensed BS
R_OWC_A = 21  # Operating Working Capital Assets (SUMIF total)
R_OWC_A_D1 = 22  # detail: AR
R_OWC_A_D2 = 23  # detail: Inventories
R_OWC_A_D3 = 24  # detail: Prepaid
R_OWC_L = 25  # Operating Working Capital Liabilities
R_OWC_L_D1 = 26  # detail: AP
R_OWC_L_D2 = 27  # detail: Accrued
R_OWC_L_D3 = 28  # detail: Deferred Rev
R_NOWC = 29
R_OLT_A = 31  # Operating Long-Term Assets
R_OLT_A_D1 = 32  # PP&E
R_OLT_A_D2 = 33  # ROU
R_OLT_A_D3 = 34  # Goodwill
R_OLT_A_D4 = 35  # Intangibles
R_OLT_A_D5 = 36  # Deferred Tax
R_OLT_A_D6 = 37  # Other LT
R_OLT_L = 38  # Operating Long-Term Liabilities
R_OLT_L_D1 = 39  # Other LT Liab
R_NOLA = 40  # Net Operating Long-Term Assets
R_NOA = 42  # Total NOA
R_FIN_L = 44  # Financial Liabilities
R_FIN_L_D1 = 45  # ST Debt
R_FIN_L_D2 = 46  # LT Debt
R_FIN_L_D3 = 47  # Lease Liab
R_FIN_A = 48  # Financial Assets
R_FIN_A_D1 = 49  # Cash
R_FIN_A_D2 = 50  # ST Investments
R_NET_DEBT = 51
R_EQUITY = 53  # Implied Equity
R_REP_EQ = 54  # Reported Equity
R_RESIDUAL = 55  # Unclassified Residual
R_TC = 56  # Total Capital
R_CHECK = 58  # NOA = TC check

# Labels
cond_labels = {
    R_OWC_A: 'Operating Working Capital Assets',
    R_OWC_A_D1: '  Accounts Receivable',
    R_OWC_A_D2: '  Inventories',
    R_OWC_A_D3: '  Prepaid & Other Current',
    R_OWC_L: 'Operating Working Capital Liabilities',
    R_OWC_L_D1: '  Accounts Payable',
    R_OWC_L_D2: '  Accrued & Current Liabilities',
    R_OWC_L_D3: '  Deferred Revenue (Current)',
    R_NOWC: 'Net Operating Working Capital (NOWC)',
    R_OLT_A: 'Operating Long-Term Assets',
    R_OLT_A_D1: '  PP&E',
    R_OLT_A_D2: '  ROU Assets (Operating)',
    R_OLT_A_D3: '  Goodwill',
    R_OLT_A_D4: '  Intangible Assets',
    R_OLT_A_D5: '  Deferred Tax Assets',
    R_OLT_A_D6: '  Other Long-Term Assets',
    R_OLT_L: 'Operating Long-Term Liabilities',
    R_OLT_L_D1: '  Other Long-Term Liabilities',
    R_NOLA: 'Net Operating Long-Term Assets',
    R_NOA: 'Total Net Operating Assets (NOA)',
    R_FIN_L: 'Financial Liabilities',
    R_FIN_L_D1: '  Short-Term Debt',
    R_FIN_L_D2: '  Long-Term Debt',
    R_FIN_L_D3: '  Lease Liabilities (Financial)',
    R_FIN_A: 'Financial Assets',
    R_FIN_A_D1: '  Cash & Equivalents',
    R_FIN_A_D2: '  Short-Term Investments',
    R_NET_DEBT: 'Net Debt (Fin Liab - Fin Assets)',
    R_EQUITY: 'Equity (Implied from Classified Items)',
    R_REP_EQ: '  Reported Equity',
    R_RESIDUAL: '  Unclassified Residual',
    R_TC: 'Total Capital (Net Debt + Equity)',
    R_CHECK: 'CHECK: NOA = Total Capital',
}
for r, lbl in cond_labels.items():
    ws_cond.cell(row=r, column=1, value=lbl).font = Font(name='Arial')

# Bold key rows
for r in [R_NOWC, R_NOLA, R_NOA, R_NET_DEBT, R_EQUITY, R_TC, R_CHECK]:
    ws_cond.cell(row=r, column=1).font = BOLD_FONT

# Build formulas for each year
for i, yr in enumerate(YEARS):
    cc = get_column_letter(i + 3)  # C, D, E, F
    dc = BS_COLS[yr]  # B, C, D, E in BS tab
    col_num = i + 3

    # SUMIF-based aggregation totals
    # Operating WC Assets
    f = f'=SUMIF($B${CLASS_START}:$B${CLASS_END},"Op. Working Capital Asset*",{cc}${CLASS_START}:{cc}${CLASS_END})'
    c = ws_cond.cell(row=R_OWC_A, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Detail rows (direct BS links for reference)
    for detail_row, bs_row in [(R_OWC_A_D1, 10), (R_OWC_A_D2, 11), (R_OWC_A_D3, 12)]:
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=detail_row, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Operating WC Liabilities
    f = f'=SUMIF($B${CLASS_START}:$B${CLASS_END},"Op. Working Capital Liability*",{cc}${CLASS_START}:{cc}${CLASS_END})'
    c = ws_cond.cell(row=R_OWC_L, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    for detail_row, bs_row in [(R_OWC_L_D1, 24), (R_OWC_L_D2, 25), (R_OWC_L_D3, 27)]:
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=detail_row, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # NOWC
    f = f"={cc}{R_OWC_A}-{cc}{R_OWC_L}"
    c = ws_cond.cell(row=R_NOWC, column=col_num); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Operating Long-Term Assets
    f = f'=SUMIF($B${CLASS_START}:$B${CLASS_END},"Op. Long-Term Asset*",{cc}${CLASS_START}:{cc}${CLASS_END})'
    c = ws_cond.cell(row=R_OLT_A, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    for detail_row, bs_row in [(R_OLT_A_D1, 15), (R_OLT_A_D2, 16), (R_OLT_A_D3, 17),
                                (R_OLT_A_D4, 18), (R_OLT_A_D5, 19), (R_OLT_A_D6, 20)]:
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=detail_row, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Operating Long-Term Liabilities
    f = f'=SUMIF($B${CLASS_START}:$B${CLASS_END},"Op. Long-Term Liability*",{cc}${CLASS_START}:{cc}${CLASS_END})'
    c = ws_cond.cell(row=R_OLT_L, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    for detail_row, bs_row in [(R_OLT_L_D1, 32)]:
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=detail_row, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # NOLA
    f = f"={cc}{R_OLT_A}-{cc}{R_OLT_L}"
    c = ws_cond.cell(row=R_NOLA, column=col_num); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Total NOA
    f = f"={cc}{R_NOWC}+{cc}{R_NOLA}"
    c = ws_cond.cell(row=R_NOA, column=col_num); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Financial Liabilities
    f = f'=SUMIF($B${CLASS_START}:$B${CLASS_END},"Financial Liability*",{cc}${CLASS_START}:{cc}${CLASS_END})'
    c = ws_cond.cell(row=R_FIN_L, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    for detail_row, bs_row in [(R_FIN_L_D1, 26), (R_FIN_L_D2, 30), (R_FIN_L_D3, 31)]:
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=detail_row, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Financial Assets
    f = f'=SUMIF($B${CLASS_START}:$B${CLASS_END},"Financial Asset*",{cc}${CLASS_START}:{cc}${CLASS_END})'
    c = ws_cond.cell(row=R_FIN_A, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = NUM_FMT

    for detail_row, bs_row in [(R_FIN_A_D1, 8), (R_FIN_A_D2, 9)]:
        f = f"='Balance Sheet'!{dc}{bs_row}"
        c = ws_cond.cell(row=detail_row, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Net Debt
    f = f"={cc}{R_FIN_L}-{cc}{R_FIN_A}"
    c = ws_cond.cell(row=R_NET_DEBT, column=col_num); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Equity (Implied) = (OWC_A + OLT_A + FIN_A) - (OWC_L + OLT_L + FIN_L)
    f = f"=({cc}{R_OWC_A}+{cc}{R_OLT_A}+{cc}{R_FIN_A})-({cc}{R_OWC_L}+{cc}{R_OLT_L}+{cc}{R_FIN_L})"
    c = ws_cond.cell(row=R_EQUITY, column=col_num); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Reported Equity
    f = f"='Balance Sheet'!{dc}39"
    c = ws_cond.cell(row=R_REP_EQ, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Residual
    f = f"={cc}{R_EQUITY}-{cc}{R_REP_EQ}"
    c = ws_cond.cell(row=R_RESIDUAL, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = NUM_FMT

    # Total Capital
    f = f"={cc}{R_NET_DEBT}+{cc}{R_EQUITY}"
    c = ws_cond.cell(row=R_TC, column=col_num); c.value = f; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Check
    f = f'=IF(ABS({cc}{R_NOA}-{cc}{R_TC})<1,"OK","CHECK")'
    c = ws_cond.cell(row=R_CHECK, column=col_num); c.value = f; c.font = BOLD_FONT; c.fill = CHECK_FILL

# ════════════════════════════════════════════════════════════════════
# TAB 5: ALT DUPONT
# ════════════════════════════════════════════════════════════════════
ws_dp = wb.create_sheet('ALT DuPont')
set_col_widths(ws_dp, {'A': 42, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 16})

ws_dp['A1'] = 'NVIDIA Corporation — Alternative DuPont Decomposition'
ws_dp['A1'].font = BOLD_FONT
ws_dp['A2'] = 'ROE = RNOA + (FLEV × Spread)'
ws_dp['A2'].font = Font(name='Arial', italic=True)
ws_dp['A3'] = 'All values linked via formulas to Condensed Financials tab'
ws_dp['A3'].font = Font(name='Arial', italic=True, color='666666')

DP_COLS = ['B', 'C', 'D', 'E', 'F']  # FY2022-FY2026
for i, dt in enumerate(FY_END_DATES):
    c = ws_dp.cell(row=5, column=i+2, value=dt)
    c.font = BOLD_FONT; c.alignment = Alignment(horizontal='center'); c.fill = HEADER_FILL
    c.number_format = DATE_FMT

# CF tab references: data is in cols C-F of Condensed Financials
CF_COLS = {'FY2022': 'C', 'FY2023': 'D', 'FY2024': 'E', 'FY2025': 'F', 'FY2026': 'G'}

# DuPont row layout
DP_R_SEC1 = 7   # PROFITABILITY section
DP_R_REV = 9
DP_R_GROW = 10
DP_R_NOPAT = 11
DP_R_MARGIN = 12
DP_R_SEC2 = 14  # EFFICIENCY
DP_R_NOA = 16
DP_R_AVG_NOA = 17
DP_R_AT = 18
DP_R_SEC3 = 20  # RNOA
DP_R_RNOA = 22
DP_R_SEC4 = 24  # FINANCIAL LEVERAGE
DP_R_ND = 26
DP_R_AVG_ND = 27
DP_R_NIAT_DP = 28
DP_R_COD = 29
DP_R_SPREAD = 30
DP_R_EQ = 32
DP_R_AVG_EQ = 33
DP_R_FLEV = 34
DP_R_FLEV_GAIN = 35
DP_R_SEC5 = 37  # ROE DECOMPOSITION
DP_R_ROE_DEC = 39
DP_R_ROE_ACT = 40
DP_R_CHECK_DP = 42
DP_R_NOTE1 = 44
DP_R_NOTE2 = 45

# Section headers
for r, lbl in [(DP_R_SEC1, 'PROFITABILITY'), (DP_R_SEC2, 'EFFICIENCY'),
               (DP_R_SEC3, 'RETURN ON NET OPERATING ASSETS'),
               (DP_R_SEC4, 'FINANCIAL LEVERAGE'), (DP_R_SEC5, 'ROE DECOMPOSITION')]:
    ws_dp.cell(row=r, column=1, value=lbl).font = BOLD_FONT
    ws_dp.cell(row=r, column=1).fill = SECTION_FILL

# Labels
dp_labels = {
    DP_R_REV: 'Revenue', DP_R_GROW: 'Sales Growth', DP_R_NOPAT: 'NOPAT',
    DP_R_MARGIN: 'NOPAT Margin', DP_R_NOA: 'Net Operating Assets (NOA)',
    DP_R_AVG_NOA: 'Average NOA', DP_R_AT: 'Asset Turnover (Rev / Avg NOA)',
    DP_R_RNOA: 'RNOA (Margin × Turnover)', DP_R_ND: 'Net Debt',
    DP_R_AVG_ND: 'Average Net Debt', DP_R_NIAT_DP: 'Net Interest After Tax',
    DP_R_COD: 'After-tax Cost of Debt', DP_R_SPREAD: 'Spread (RNOA - Cost of Debt)',
    DP_R_EQ: 'Equity (Implied)', DP_R_AVG_EQ: 'Average Equity',
    DP_R_FLEV: 'Net Financial Leverage (Avg ND / Avg Eq)',
    DP_R_FLEV_GAIN: 'Financial Leverage Gain (Spread × FLEV)',
    DP_R_ROE_DEC: 'ROE Decomposed (RNOA + FLEV Gain)',
    DP_R_ROE_ACT: 'ROE Actual (NI / Avg Implied Equity)',
    DP_R_CHECK_DP: 'CHECK: |Decomposed - Actual| < 0.01%',
}
for r, lbl in dp_labels.items():
    ws_dp.cell(row=r, column=1, value=lbl).font = Font(name='Arial')

for r in [DP_R_RNOA, DP_R_ROE_DEC, DP_R_ROE_ACT, DP_R_CHECK_DP]:
    ws_dp.cell(row=r, column=1).font = BOLD_FONT

# Build DuPont formulas
for i, yr in enumerate(YEARS):
    dp_col = get_column_letter(i + 2)  # B, C, D, E
    cf_col = CF_COLS[yr]  # C, D, E, F in Condensed Financials
    is_col = IS_COLS[yr]  # B, C, D, E in IS tab
    col_num = i + 2
    is_first = (i == 0)

    # Revenue (from IS)
    f = f"='Income Statement'!{is_col}7"
    c = ws_dp.cell(row=DP_R_REV, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Sales Growth
    if is_first:
        ws_dp.cell(row=DP_R_GROW, column=col_num, value='n/a').font = GRAY_FONT
    else:
        prev = get_column_letter(col_num - 1)
        f = f"={dp_col}{DP_R_REV}/{prev}{DP_R_REV}-1"
        c = ws_dp.cell(row=DP_R_GROW, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # NOPAT (from CF tab)
    f = f"='Condensed Financials'!{cf_col}{R_NOPAT}"
    c = ws_dp.cell(row=DP_R_NOPAT, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # NOPAT Margin
    f = f"=IFERROR({dp_col}{DP_R_NOPAT}/{dp_col}{DP_R_REV},0)"
    c = ws_dp.cell(row=DP_R_MARGIN, column=col_num); c.value = f; c.font = BLACK_FONT; c.number_format = PCT_FMT

    # NOA (from CF tab)
    f = f"='Condensed Financials'!{cf_col}{R_NOA}"
    c = ws_dp.cell(row=DP_R_NOA, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Avg NOA
    if is_first:
        ws_dp.cell(row=DP_R_AVG_NOA, column=col_num, value='n/a').font = GRAY_FONT
    else:
        prev = get_column_letter(col_num - 1)
        f = f"=({dp_col}{DP_R_NOA}+{prev}{DP_R_NOA})/2"
        c = ws_dp.cell(row=DP_R_AVG_NOA, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = NUM_FMT

    # Asset Turnover
    if is_first:
        ws_dp.cell(row=DP_R_AT, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"=IFERROR({dp_col}{DP_R_REV}/{dp_col}{DP_R_AVG_NOA},0)"
        c = ws_dp.cell(row=DP_R_AT, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = '0.00x'

    # RNOA
    if is_first:
        ws_dp.cell(row=DP_R_RNOA, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"={dp_col}{DP_R_MARGIN}*{dp_col}{DP_R_AT}"
        c = ws_dp.cell(row=DP_R_RNOA, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # Net Debt
    f = f"='Condensed Financials'!{cf_col}{R_NET_DEBT}"
    c = ws_dp.cell(row=DP_R_ND, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Avg Net Debt
    if is_first:
        ws_dp.cell(row=DP_R_AVG_ND, column=col_num, value='n/a').font = GRAY_FONT
    else:
        prev = get_column_letter(col_num - 1)
        f = f"=({dp_col}{DP_R_ND}+{prev}{DP_R_ND})/2"
        c = ws_dp.cell(row=DP_R_AVG_ND, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = NUM_FMT

    # Net Interest After Tax
    f = f"='Condensed Financials'!{cf_col}{R_NIAT}"
    c = ws_dp.cell(row=DP_R_NIAT_DP, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # After-tax Cost of Debt
    if is_first:
        ws_dp.cell(row=DP_R_COD, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"=IFERROR({dp_col}{DP_R_NIAT_DP}/{dp_col}{DP_R_AVG_ND},0)"
        c = ws_dp.cell(row=DP_R_COD, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # Spread
    if is_first:
        ws_dp.cell(row=DP_R_SPREAD, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"={dp_col}{DP_R_RNOA}-{dp_col}{DP_R_COD}"
        c = ws_dp.cell(row=DP_R_SPREAD, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # Equity
    f = f"='Condensed Financials'!{cf_col}{R_EQUITY}"
    c = ws_dp.cell(row=DP_R_EQ, column=col_num); c.value = f; c.font = GREEN_FONT; c.number_format = NUM_FMT

    # Avg Equity
    if is_first:
        ws_dp.cell(row=DP_R_AVG_EQ, column=col_num, value='n/a').font = GRAY_FONT
    else:
        prev = get_column_letter(col_num - 1)
        f = f"=({dp_col}{DP_R_EQ}+{prev}{DP_R_EQ})/2"
        c = ws_dp.cell(row=DP_R_AVG_EQ, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = NUM_FMT

    # FLEV
    if is_first:
        ws_dp.cell(row=DP_R_FLEV, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"=IFERROR({dp_col}{DP_R_AVG_ND}/{dp_col}{DP_R_AVG_EQ},0)"
        c = ws_dp.cell(row=DP_R_FLEV, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = '0.00x'

    # FLEV Gain
    if is_first:
        ws_dp.cell(row=DP_R_FLEV_GAIN, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"={dp_col}{DP_R_SPREAD}*{dp_col}{DP_R_FLEV}"
        c = ws_dp.cell(row=DP_R_FLEV_GAIN, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # ROE Decomposed
    if is_first:
        ws_dp.cell(row=DP_R_ROE_DEC, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"={dp_col}{DP_R_RNOA}+{dp_col}{DP_R_FLEV_GAIN}"
        c = ws_dp.cell(row=DP_R_ROE_DEC, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # ROE Actual
    if is_first:
        ws_dp.cell(row=DP_R_ROE_ACT, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f"=IFERROR('Income Statement'!{is_col}21/{dp_col}{DP_R_AVG_EQ},0)"
        c = ws_dp.cell(row=DP_R_ROE_ACT, column=col_num); c.value = f; c.font = GRAY_FONT; c.number_format = PCT_FMT

    # Check
    if is_first:
        ws_dp.cell(row=DP_R_CHECK_DP, column=col_num, value='n/a').font = GRAY_FONT
    else:
        f = f'=IF(ABS({dp_col}{DP_R_ROE_DEC}-{dp_col}{DP_R_ROE_ACT})<0.0001,"OK","Off by "&TEXT(ABS({dp_col}{DP_R_ROE_DEC}-{dp_col}{DP_R_ROE_ACT}),"0.00%"))'
        c = ws_dp.cell(row=DP_R_CHECK_DP, column=col_num); c.value = f; c.font = BOLD_FONT; c.fill = CHECK_FILL

# Notes
ws_dp.cell(row=DP_R_NOTE1, column=1, value='NOTE: NVIDIA has negative Net Debt (cash > debt).').font = Font(name='Arial', italic=True, color='666666')
ws_dp.cell(row=DP_R_NOTE2, column=1, value='FLEV is negative → leverage REDUCES ROE vs RNOA.').font = Font(name='Arial', italic=True, color='666666')

# ════════════════════════════════════════════════════════════════════
# TAB 6-8: MODEL TABS (Bull, Base, Bear)
# Following reference script row layout exactly
# ════════════════════════════════════════════════════════════════════

import json

with open('/sessions/zen-zealous-hamilton/NVDA_Assumptions.json') as f:
    assumptions = json.load(f)

RF = assumptions['market']['risk_free_rate']
ERP = assumptions['market']['equity_risk_premium']
SHARES = assumptions['market']['shares_outstanding_M']
PRICE = assumptions['market']['current_price']
TAX_RATE = 0.21  # NVDA effective blended rate; user can override

# Scenario configs from assumptions JSON
scenarios = {}
for sname in ['Bull', 'Base', 'Bear']:
    s = assumptions['scenarios'][sname]
    # Support both old format (nowc_sales scalar, nola_sales scalar)
    # and new format (nowc_vector, nola_vector)
    if 'nowc_vector' in s:
        nowc_vec = s['nowc_vector']
    else:
        nowc_vec = [s['nowc_sales']] * 10
    if 'nola_vector' in s:
        nola_vec = s['nola_vector']
    else:
        nola_vec = [s['nola_sales']] * 10
    term_nowc = nowc_vec[-1]
    term_nola = nola_vec[-1]
    term_margin = s['margin_vector'][-1]
    term_noa_sales = term_nowc + term_nola
    term_rnoa = term_margin / term_noa_sales if term_noa_sales > 0 else 0
    scenarios[sname] = {
        'beta': s['beta'],
        'probability': s['probability'],
        'growthVector': s['growth_vector'],
        'marginVector': s['margin_vector'],
        'nowcRatioVector': nowc_vec,
        'nolaRatioVector': nola_vec,
        'notes': {
            'growthRationale': s['narrative'],
            'moatRationale': (
                f"NOPAT margins fade from {s['margin_vector'][0]:.0%} to {term_margin:.0%}. "
                f"Asset intensity rises (NOA/Sales: {nowc_vec[0]+nola_vec[0]:.2f} → {term_noa_sales:.2f}). "
                f"Terminal RNOA: {term_rnoa:.0%} vs CoE {s['cost_of_equity']:.1%} ({term_rnoa/s['cost_of_equity']:.1f}x)."
            ),
            'keyRisks': 'See Scenario Summary for full risk analysis.',
            'valuationContext': f"Probability weight: {s['probability']:.0%}. Compare IVPS to current price ${PRICE}."
        }
    }

# ── Column layout: Historical years + 10 forecast years ──
# Historical: FY2022-FY2025 in cols B-E; Forecast: Y1-Y10 in cols F-O
NUM_HIST = len(YEARS)
LAST_CF_COL = get_column_letter(NUM_HIST + 2)  # For 5 years: column G  # 4
PROJ_YEARS = 10
HIST_COLS = [get_column_letter(c) for c in range(2, 2 + NUM_HIST)]      # B,C,D,E
FC_COLS = [get_column_letter(c) for c in range(2 + NUM_HIST, 2 + NUM_HIST + PROJ_YEARS)]  # F,G,...,O
FIRST_FC = FC_COLS[0]   # F
LAST_FC = FC_COLS[-1]    # O
LAST_HIST = HIST_COLS[-1]  # E
ALL_COLS = HIST_COLS + FC_COLS  # B through O

# Condensed Financials uses cols C-F for FY2022-FY2025 (offset +1 from IS/BS)
# So hist year i → CF col = get_column_letter(i+3)
# ALT DuPont uses cols B-E for FY2022-FY2025 (same as IS/BS)

# Source row references in other tabs
IS_SALES_ROW = 7        # Revenue row in Income Statement
IS_NI_ROW = 21           # Net Income row in IS
CF_NOPAT_ROW = R_NOPAT   # 17 in Condensed Financials
CF_NIAT_ROW = R_NIAT     # 16
CF_NOWC_ROW = R_NOWC     # 29
CF_NOLA_ROW = R_NOLA     # 40
CF_NOA_ROW = R_NOA       # 42
CF_ND_ROW = R_NET_DEBT   # 51
CF_EQ_ROW = R_EQUITY     # 53
DP_COD_ROW = DP_R_COD    # 29 in ALT DuPont

# Formatting fills
ORANGE_FILL = PatternFill('solid', fgColor='FCE5CD')  # assumptions
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')   # returns
GREEN_FILL_LIGHT = PatternFill('solid', fgColor='D9EAD3')  # AE/FCF

for scenario_name, s in scenarios.items():
    ws_m = wb.create_sheet(f'Model_{scenario_name}')
    widths = {'A': 48}
    for col_letter in ALL_COLS:
        widths[col_letter] = 14
    set_col_widths(ws_m, widths)

    g_vec = s['growthVector']
    m_vec = s['marginVector']
    nowc_vec = s['nowcRatioVector']
    nola_vec = s['nolaRatioVector']

    # ── ROWS 1-8: MARKET DATA ──
    ws_m.cell(row=1, column=1, value='Market Data (Given)').font = BOLD_FONT
    for r, lbl, val, fmt in [
        (2, 'Beta', s['beta'], '0.00'),
        (3, 'Risk premium', ERP, PCT_FMT),
        (4, 'Risk free rate', RF, PCT_FMT),
        (6, 'Tax Rate', TAX_RATE, PCT_FMT),
    ]:
        ws_m.cell(row=r, column=1, value=lbl).font = Font(name='Arial')
        c = ws_m.cell(row=r, column=2, value=val)
        c.font = BLUE_FONT; c.number_format = fmt

    # Row 5: Cost of equity = Rf + ERP × Beta (FORMULA)
    ws_m.cell(row=5, column=1, value='Cost of equity').font = Font(name='Arial')
    c = ws_m.cell(row=5, column=2); c.value = '=B4+B3*B2'; c.font = BLACK_FONT; c.number_format = PCT_FMT

    # Row 7: Pre-tax cost of debt = AVERAGE of ALT DuPont historical cost of debt
    # ALT DuPont cost of debt is in row DP_COD_ROW, cols C,D,E (FY2023,FY2024,FY2025 — skip FY2022 which is n/a)
    ws_m.cell(row=7, column=1, value='Pre-Tax Net Cost of Net Debt').font = Font(name='Arial')
    c = ws_m.cell(row=7, column=2)
    c.value = f"=AVERAGE('ALT DuPont'!C{DP_COD_ROW}:F{DP_COD_ROW})"
    c.font = BLACK_FONT; c.number_format = PCT_FMT

    # Row 8: After-tax cost of debt = Pre-tax × (1 - Tax Rate)
    ws_m.cell(row=8, column=1, value='After-Tax Net Cost of Net Debt').font = Font(name='Arial')
    c = ws_m.cell(row=8, column=2); c.value = '=B7*(1-B6)'; c.font = BLACK_FONT; c.number_format = PCT_FMT

    # ── ROWS 10-17: VALUATION ASSUMPTIONS ──
    # Row 10: Year headers — use YEAR() on IS date headers for historical, +1 for forecast
    for i in range(NUM_HIST):
        is_col = DATA_COLS[i]  # B,C,D,E in IS tab
        col_num = i + 2
        c = ws_m.cell(row=10, column=col_num)
        c.value = f"=YEAR('Income Statement'!{is_col}6)"
        c.font = BOLD_FONT; c.alignment = Alignment(horizontal='center')
        c.number_format = '0'
    for j in range(PROJ_YEARS):
        col_num = NUM_HIST + 2 + j
        prev_col_l = HIST_COLS[-1] if j == 0 else FC_COLS[j-1]
        c = ws_m.cell(row=10, column=col_num)
        c.value = f"={prev_col_l}10+1"
        c.font = BOLD_FONT; c.alignment = Alignment(horizontal='center')
        c.number_format = '0'

    # Row 11: Revenue Growth
    ws_m.cell(row=11, column=1, value='Revenue Growth Assumption').font = Font(name='Arial')
    # Historical: calculated from actual Sales (skip first year — no prior)
    for i in range(1, NUM_HIST):
        col_l = HIST_COLS[i]
        prev_l = HIST_COLS[i-1]
        col_num = i + 2
        c = ws_m.cell(row=11, column=col_num)
        c.value = f"={col_l}31/{prev_l}31-1"; c.font = BLACK_FONT; c.number_format = PCT_FMT
        c.fill = ORANGE_FILL
    # Forecast: hardcoded from vector
    for j in range(PROJ_YEARS):
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=11, column=col_num, value=round(g_vec[j], 4))
        c.font = BLUE_FONT; c.number_format = PCT_FMT; c.fill = ORANGE_FILL

    # Row 12: NOPAT Margin
    ws_m.cell(row=12, column=1, value='NOPAT Margin').font = Font(name='Arial')
    for i in range(NUM_HIST):
        col_l = HIST_COLS[i]
        col_num = i + 2
        c = ws_m.cell(row=12, column=col_num)
        c.value = f"={col_l}32/{col_l}31"; c.font = BLACK_FONT; c.number_format = PCT_FMT
        c.fill = ORANGE_FILL
    for j in range(PROJ_YEARS):
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=12, column=col_num, value=round(m_vec[j], 4))
        c.font = BLUE_FONT; c.number_format = PCT_FMT; c.fill = ORANGE_FILL

    # Row 14: NOWC / Sales
    ws_m.cell(row=14, column=1, value='(Beg) Net Operating Working Capital / Sales').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_l = HIST_COLS[i]
        col_num = i + 2
        c = ws_m.cell(row=14, column=col_num)
        c.value = f"={col_l}21/{col_l}31"; c.font = BLACK_FONT; c.number_format = '0.000'
        c.fill = ORANGE_FILL
    # First forecast: calculated from anchor
    c = ws_m.cell(row=14, column=NUM_HIST + 2)
    c.value = f"={FIRST_FC}21/{FIRST_FC}31"; c.font = BLACK_FONT; c.number_format = '0.000'
    c.fill = ORANGE_FILL
    # Remaining forecast: hardcoded from vector (increasing over time)
    for j in range(1, PROJ_YEARS):
        col_num = NUM_HIST + 2 + j
        val = nowc_vec[j] if j < len(nowc_vec) else nowc_vec[-1]
        c = ws_m.cell(row=14, column=col_num, value=round(val, 4))
        c.font = BLUE_FONT; c.number_format = '0.000'; c.fill = ORANGE_FILL

    # Row 15: NOLA / Sales
    ws_m.cell(row=15, column=1, value='(Beg) Net Operating LT Assets / Sales').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_l = HIST_COLS[i]
        col_num = i + 2
        c = ws_m.cell(row=15, column=col_num)
        c.value = f"={col_l}22/{col_l}31"; c.font = BLACK_FONT; c.number_format = '0.000'
        c.fill = ORANGE_FILL
    # First forecast: calculated
    c = ws_m.cell(row=15, column=NUM_HIST + 2)
    c.value = f"={FIRST_FC}22/{FIRST_FC}31"; c.font = BLACK_FONT; c.number_format = '0.000'
    c.fill = ORANGE_FILL
    # Remaining: vector
    for j in range(1, PROJ_YEARS):
        col_num = NUM_HIST + 2 + j
        val = nola_vec[j] if j < len(nola_vec) else nola_vec[-1]
        c = ws_m.cell(row=15, column=col_num, value=round(val, 4))
        c.font = BLUE_FONT; c.number_format = '0.000'; c.fill = ORANGE_FILL

    # Row 17: Leverage (Net Debt / Total Capital)
    ws_m.cell(row=17, column=1, value='Leverage (Beg. Net debt / Beg. Assets)').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_l = HIST_COLS[i]
        col_num = i + 2
        c = ws_m.cell(row=17, column=col_num)
        c.value = f"={col_l}26/{col_l}28"; c.font = BLACK_FONT; c.number_format = '0.000'
        c.fill = ORANGE_FILL
    # First forecast: calculated
    c = ws_m.cell(row=17, column=NUM_HIST + 2)
    c.value = f"={FIRST_FC}26/{FIRST_FC}28"; c.font = BLACK_FONT; c.number_format = '0.000'
    c.fill = ORANGE_FILL
    # Remaining: carry forward
    for j in range(1, PROJ_YEARS):
        col_l = FC_COLS[j]
        prev_l = FC_COLS[j-1]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=17, column=col_num)
        c.value = f"={prev_l}17"; c.font = BLACK_FONT; c.number_format = '0.000'
        c.fill = ORANGE_FILL

    # ── ROWS 19-28: BALANCE SHEET (Beginning) ──
    ws_m.cell(row=19, column=1, value='BALANCE SHEET (Beginning)').font = BOLD_FONT
    # Year headers (same as row 10)
    for col_num in range(2, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=19, column=col_num)
        c.value = f"={col_l}10"; c.font = BOLD_FONT; c.alignment = Alignment(horizontal='center')

    ws_m.cell(row=20, column=1, value='Operating Assets (Beginning)').font = Font(name='Arial', bold=True)
    ws_m.cell(row=25, column=1, value='Capital (Beginning)').font = Font(name='Arial', bold=True)

    # Row 21: NOWC
    ws_m.cell(row=21, column=1, value='  Net operating working capital').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_num = i + 2
        cf_col = get_column_letter(i + 3)  # CF uses cols C-F
        c = ws_m.cell(row=21, column=col_num)
        c.value = f"='Condensed Financials'!{cf_col}{CF_NOWC_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    # First forecast: link to CF anchor
    cf_anchor = get_column_letter(NUM_HIST + 2)  # F in CF = col for last year + 1? No...
    # CF data cols are C,D,E,F for FY2022-FY2025. Last = F.
    # First forecast anchor = last historical CF value = CF col F (for FY2025)
    c = ws_m.cell(row=21, column=NUM_HIST + 2)
    c.value = f"='Condensed Financials'!{LAST_CF_COL}{CF_NOWC_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    # Remaining forecast: NOWC/Sales ratio × Sales
    for j in range(1, PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=21, column=col_num)
        c.value = f"={col_l}14*{col_l}$31"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 22: NOLA
    ws_m.cell(row=22, column=1, value='  Net operating LT assets').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_num = i + 2
        cf_col = get_column_letter(i + 3)
        c = ws_m.cell(row=22, column=col_num)
        c.value = f"='Condensed Financials'!{cf_col}{CF_NOLA_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    c = ws_m.cell(row=22, column=NUM_HIST + 2)
    c.value = f"='Condensed Financials'!{LAST_CF_COL}{CF_NOLA_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    for j in range(1, PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=22, column=col_num)
        c.value = f"={col_l}15*{col_l}$31"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 23: Total NOA
    ws_m.cell(row=23, column=1, value='    Total operating net assets').font = Font(name='Arial')
    for col_num in range(3, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=23, column=col_num)
        c.value = f"={col_l}21+{col_l}22"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 26: Net Debt
    ws_m.cell(row=26, column=1, value='  Net debt').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_num = i + 2
        cf_col = get_column_letter(i + 3)
        c = ws_m.cell(row=26, column=col_num)
        c.value = f"='Condensed Financials'!{cf_col}{CF_ND_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    c = ws_m.cell(row=26, column=NUM_HIST + 2)
    c.value = f"='Condensed Financials'!{LAST_CF_COL}{CF_ND_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    for j in range(1, PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=26, column=col_num)
        c.value = f"={col_l}17*{col_l}23"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 27: Equity
    ws_m.cell(row=27, column=1, value='  Common equity').font = Font(name='Arial')
    for i in range(1, NUM_HIST):
        col_num = i + 2
        cf_col = get_column_letter(i + 3)
        c = ws_m.cell(row=27, column=col_num)
        c.value = f"='Condensed Financials'!{cf_col}{CF_EQ_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    c = ws_m.cell(row=27, column=NUM_HIST + 2)
    c.value = f"='Condensed Financials'!{LAST_CF_COL}{CF_EQ_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    for j in range(1, PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=27, column=col_num)
        c.value = f"={col_l}23-{col_l}26"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 28: Total Capital
    ws_m.cell(row=28, column=1, value='    Total Capital').font = Font(name='Arial')
    for col_num in range(3, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=28, column=col_num)
        c.value = f"={col_l}26+{col_l}27"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # ── ROWS 30-37: INCOME STATEMENT ──
    ws_m.cell(row=30, column=1, value='INCOME STATEMENT').font = BOLD_FONT
    # Year headers
    for col_num in range(2, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=30, column=col_num)
        c.value = f"={col_l}10"; c.font = BOLD_FONT; c.alignment = Alignment(horizontal='center')

    # Row 31: Sales
    ws_m.cell(row=31, column=1, value='Sales').font = Font(name='Arial')
    for i in range(NUM_HIST):
        is_col = DATA_COLS[i]  # B,C,D,E in IS
        col_num = i + 2
        c = ws_m.cell(row=31, column=col_num)
        c.value = f"='Income Statement'!{is_col}{IS_SALES_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        prev_l = LAST_HIST if j == 0 else FC_COLS[j-1]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=31, column=col_num)
        c.value = f"={prev_l}31*(1+{col_l}11)"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 32: NOPAT
    ws_m.cell(row=32, column=1, value='NOPAT').font = Font(name='Arial')
    for i in range(NUM_HIST):
        cf_col = get_column_letter(i + 3)  # C,D,E,F in CF
        col_num = i + 2
        c = ws_m.cell(row=32, column=col_num)
        c.value = f"='Condensed Financials'!{cf_col}{CF_NOPAT_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=32, column=col_num)
        c.value = f"={col_l}31*{col_l}12"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 33: After-tax interest
    ws_m.cell(row=33, column=1, value='After tax interest expense (income)').font = Font(name='Arial')
    for i in range(NUM_HIST):
        cf_col = get_column_letter(i + 3)
        col_num = i + 2
        c = ws_m.cell(row=33, column=col_num)
        c.value = f"='Condensed Financials'!{cf_col}{CF_NIAT_ROW}"; c.font = GREEN_FONT; c.number_format = NUM_FMT
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=33, column=col_num)
        c.value = f"={col_l}26*$B$8"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 34: Net Income
    ws_m.cell(row=34, column=1, value='Net Income').font = Font(name='Arial')
    for col_num in range(2, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=34, column=col_num)
        c.value = f"={col_l}32-{col_l}33"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 36: Return on Equity (yellow background, bold)
    ws_m.cell(row=36, column=1, value='Return on (Beg) Equity').font = BOLD_FONT
    for col_num in range(3, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=36, column=col_num)
        c.value = f"=IFERROR({col_l}34/{col_l}27,0)"; c.font = BOLD_FONT; c.number_format = PCT_FMT
        c.fill = YELLOW_FILL

    # Row 37: Return on NOA (yellow background, bold)
    ws_m.cell(row=37, column=1, value='Return on (Beg) NOA').font = BOLD_FONT
    for col_num in range(3, 2 + NUM_HIST + PROJ_YEARS):
        col_l = get_column_letter(col_num)
        c = ws_m.cell(row=37, column=col_num)
        c.value = f"=IFERROR({col_l}32/{col_l}23,0)"; c.font = BOLD_FONT; c.number_format = PCT_FMT
        c.fill = YELLOW_FILL

    # ── ROWS 39-51: ABNORMAL EARNINGS & VALUATION ──
    ws_m.cell(row=39, column=1, value='Abnormal Earnings').font = BOLD_FONT
    # Year headers (forecast only)
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=39, column=col_num)
        c.value = f"={col_l}10"; c.font = BOLD_FONT; c.alignment = Alignment(horizontal='center')

    # Row 40: Net Income (link)
    ws_m.cell(row=40, column=1, value='Net Income').font = Font(name='Arial')
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=40, column=col_num)
        c.value = f"={col_l}34"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 41: Capital Charge
    ws_m.cell(row=41, column=1, value='Capital charge').font = Font(name='Arial')
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=41, column=col_num)
        c.value = f"={col_l}27*$B$5"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 42: Abnormal Earnings (green background)
    ws_m.cell(row=42, column=1, value='Abnormal earnings').font = BOLD_FONT
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=42, column=col_num)
        c.value = f"={col_l}40-{col_l}41"; c.font = BOLD_FONT; c.number_format = NUM_FMT
        c.fill = GREEN_FILL_LIGHT

    # Row 43: FCF to Equity (green background)
    ws_m.cell(row=43, column=1, value='Free Cash Flow to Equity').font = BOLD_FONT
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        if j == 0:
            prev_eq = LAST_HIST  # last historical equity
        else:
            prev_eq = FC_COLS[j-1]
        c = ws_m.cell(row=43, column=col_num)
        c.value = f"={col_l}40+({prev_eq}27-{col_l}27)"; c.font = BOLD_FONT; c.number_format = NUM_FMT
        c.fill = GREEN_FILL_LIGHT

    # Row 44: PV Factor
    ws_m.cell(row=44, column=1, value='PV Factor (using cost of equity)').font = Font(name='Arial')
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        if j == 0:
            c = ws_m.cell(row=44, column=col_num)
            c.value = "=1/(1+$B$5)"; c.font = BLACK_FONT; c.number_format = '0.0000'
        else:
            prev_l = FC_COLS[j-1]
            c = ws_m.cell(row=44, column=col_num)
            c.value = f"={prev_l}44/(1+$B$5)"; c.font = BLACK_FONT; c.number_format = '0.0000'

    # Row 45: PV of AE
    ws_m.cell(row=45, column=1, value='PV of abnormal earnings').font = Font(name='Arial')
    for j in range(PROJ_YEARS):
        col_l = FC_COLS[j]
        col_num = NUM_HIST + 2 + j
        c = ws_m.cell(row=45, column=col_num)
        c.value = f"={col_l}42*{col_l}44"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 46: Sum of PV of AE (first FC col only)
    # But in last FC col, we put the terminal value formula
    ws_m.cell(row=46, column=1, value='Sum of PV of abnormal earnings').font = Font(name='Arial')
    fc_first_num = NUM_HIST + 2
    c = ws_m.cell(row=46, column=fc_first_num)
    c.value = f"=SUM({FIRST_FC}45:{LAST_FC}45)"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Terminal value in last FC col row 46
    fc_last_num = NUM_HIST + 2 + PROJ_YEARS - 1
    c = ws_m.cell(row=46, column=fc_last_num)
    c.value = f"={LAST_FC}42*(1+${LAST_FC}$11)/($B$5-${LAST_FC}$11)"
    c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 47: Terminal value PV (first FC col only)
    ws_m.cell(row=47, column=1, value='Terminal value (PV)').font = Font(name='Arial')
    c = ws_m.cell(row=47, column=fc_first_num)
    c.value = f"={LAST_FC}46*{LAST_FC}44"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 48: Book Value of Equity
    ws_m.cell(row=48, column=1, value='Book Value of Equity').font = Font(name='Arial')
    c = ws_m.cell(row=48, column=fc_first_num)
    c.value = f"={FIRST_FC}27"; c.font = BLACK_FONT; c.number_format = NUM_FMT

    # Row 49: Intrinsic Value
    ws_m.cell(row=49, column=1, value='Intrinsic Value of Common Equity').font = BOLD_FONT
    c = ws_m.cell(row=49, column=fc_first_num)
    c.value = f"={FIRST_FC}48+{FIRST_FC}47+{FIRST_FC}46"; c.font = BOLD_FONT; c.number_format = NUM_FMT

    # Row 50: Shares
    ws_m.cell(row=50, column=1, value='Number of shares (diluted)').font = Font(name='Arial')
    c = ws_m.cell(row=50, column=fc_first_num, value=SHARES)
    c.font = BLUE_FONT; c.number_format = '#,##0'

    # Row 51: IVPS
    ws_m.cell(row=51, column=1, value='Intrinsic value per share').font = BOLD_FONT
    c = ws_m.cell(row=51, column=fc_first_num)
    c.value = f"={FIRST_FC}49/{FIRST_FC}50"; c.font = BOLD_FONT; c.number_format = PRICE_FMT

    # ── ROWS 55+: PROFESSOR'S NOTES ──
    notes = s['notes']
    ws_m.cell(row=55, column=1, value=f"PROFESSOR'S STRATEGIC NOTES — {scenario_name.upper()} CASE").font = BOLD_FONT
    ws_m.cell(row=57, column=1, value='Growth Assumption Rationale:').font = BOLD_FONT
    ws_m.cell(row=58, column=1, value=notes['growthRationale']).font = Font(name='Arial', italic=True)
    ws_m.cell(row=60, column=1, value='Margin & Moat Assessment:').font = BOLD_FONT
    ws_m.cell(row=61, column=1, value=notes['moatRationale']).font = Font(name='Arial', italic=True)
    ws_m.cell(row=63, column=1, value='Key Risks to Monitor:').font = BOLD_FONT
    ws_m.cell(row=64, column=1, value=notes['keyRisks']).font = Font(name='Arial', italic=True)
    ws_m.cell(row=66, column=1, value='Valuation Context:').font = BOLD_FONT
    ws_m.cell(row=67, column=1, value=notes['valuationContext']).font = Font(name='Arial', italic=True)

    # Freeze column A
    ws_m.freeze_panes = 'B1'

# ════════════════════════════════════════════════════════════════════
# TAB 9: SCENARIO SUMMARY
# ════════════════════════════════════════════════════════════════════
ws_ss = wb.create_sheet('Scenario_Summary')
set_col_widths(ws_ss, {'A': 48, 'B': 18, 'C': 18, 'D': 18, 'E': 18})

ws_ss['A1'] = 'NVIDIA Corporation (NVDA) — SCENARIO COMPARISON'
ws_ss['A1'].font = Font(name='Arial', bold=True, size=14)
ws_ss['A2'] = 'Multi-Scenario Residual Income Valuation'
ws_ss['A2'].font = Font(name='Arial', italic=True)

# Column headers
BLUE_BG = PatternFill('solid', fgColor='4A86E8')
WHITE_BOLD = Font(name='Arial', bold=True, color='FFFFFF')
for col_num, lbl in [(1, 'Metric'), (2, 'Bear'), (3, 'Base'), (4, 'Bull')]:
    c = ws_ss.cell(row=4, column=col_num, value=lbl)
    c.font = WHITE_BOLD; c.fill = BLUE_BG; c.alignment = Alignment(horizontal='center')

# Map: col 2=Bear, col 3=Base, col 4=Bull
model_tabs_ss = {2: 'Model_Bear', 3: 'Model_Base', 4: 'Model_Bull'}

# ── VALUATION OUTPUTS ──
ws_ss.cell(row=5, column=1, value='VALUATION OUTPUTS').font = BOLD_FONT

valuation_metrics = [
    (6, 'Intrinsic Value per Share', 51, PRICE_FMT),
    (7, 'Intrinsic Value of Equity ($M)', 49, NUM_FMT),
    (8, 'Terminal Value (PV, $M)', 47, NUM_FMT),
    (9, 'Sum of PV Abnormal Earnings ($M)', 46, NUM_FMT),
    (10, 'Beginning Book Equity ($M)', 48, NUM_FMT),
]
for r, lbl, src_row, fmt in valuation_metrics:
    ws_ss.cell(row=r, column=1, value=lbl).font = Font(name='Arial')
    for col_num, tab in model_tabs_ss.items():
        c = ws_ss.cell(row=r, column=col_num)
        c.value = f"='{tab}'!{FIRST_FC}{src_row}"; c.font = GREEN_FONT; c.number_format = fmt

# ── KEY ASSUMPTIONS ──
ws_ss.cell(row=12, column=1, value='KEY ASSUMPTIONS').font = BOLD_FONT

assumption_rows = [
    (13, 'Year 1 Revenue Growth', FIRST_FC, 11, PCT_FMT),
    (14, 'Terminal Revenue Growth', LAST_FC, 11, PCT_FMT),
    (15, 'Year 1 NOPAT Margin', FIRST_FC, 12, PCT_FMT),
    (16, 'Terminal NOPAT Margin', LAST_FC, 12, PCT_FMT),
    (17, 'Cost of Equity', 'B', 5, PCT_FMT),
    (18, 'Beta', 'B', 2, '0.00'),
]
for r, lbl, col_ref, src_row, fmt in assumption_rows:
    ws_ss.cell(row=r, column=1, value=lbl).font = Font(name='Arial')
    for col_num, tab in model_tabs_ss.items():
        c = ws_ss.cell(row=r, column=col_num)
        c.value = f"='{tab}'!{col_ref}{src_row}"; c.font = GREEN_FONT; c.number_format = fmt

# ── TERMINAL YEAR RETURNS ──
ws_ss.cell(row=20, column=1, value='TERMINAL YEAR RETURNS').font = BOLD_FONT

for r, lbl, src_row in [(21, 'Terminal ROE', 36), (22, 'Terminal ROIC (RNOA)', 37)]:
    ws_ss.cell(row=r, column=1, value=lbl).font = Font(name='Arial')
    for col_num, tab in model_tabs_ss.items():
        c = ws_ss.cell(row=r, column=col_num)
        c.value = f"='{tab}'!{LAST_FC}{src_row}"; c.font = GREEN_FONT; c.number_format = PCT_FMT

# ── PROBABILITY-WEIGHTED VALUATION ──
PROB_FILL = PatternFill('solid', fgColor='F4CCCC')
ws_ss.cell(row=24, column=1, value='PROBABILITY-WEIGHTED VALUATION').font = BOLD_FONT

# Row 25: Probability
ws_ss.cell(row=25, column=1, value='Scenario Probability').font = Font(name='Arial')
for col_num, sname in [(2, 'Bear'), (3, 'Base'), (4, 'Bull')]:
    prob = assumptions['scenarios'][sname]['probability']
    c = ws_ss.cell(row=25, column=col_num, value=prob)
    c.font = BLUE_FONT; c.number_format = PCT_FMT; c.fill = PROB_FILL

# Row 26: Weighted IVPS
ws_ss.cell(row=26, column=1, value='Probability-Weighted Value/Share').font = Font(name='Arial', bold=True, size=12)
c = ws_ss.cell(row=26, column=2)
c.value = f"=B6*B25+C6*C25+D6*D25"
c.font = Font(name='Arial', bold=True, size=12); c.number_format = PRICE_FMT; c.fill = PROB_FILL

# Row 27: Current Price
ws_ss.cell(row=27, column=1, value='Current Stock Price').font = Font(name='Arial')
c = ws_ss.cell(row=27, column=2, value=PRICE)
c.font = BLUE_FONT; c.number_format = PRICE_FMT; c.fill = PROB_FILL

# Row 28: Upside
ws_ss.cell(row=28, column=1, value='Implied Upside/(Downside)').font = BOLD_FONT
c = ws_ss.cell(row=28, column=2)
c.value = "=B26/B27-1"; c.font = BOLD_FONT; c.number_format = PCT_FMT; c.fill = PROB_FILL

# ── SCENARIO VALUE RANGE ──
ws_ss.cell(row=30, column=1, value='SCENARIO VALUE RANGE').font = BOLD_FONT
for r, lbl, src_col in [(31, 'Bear Case Value', 'B'), (32, 'Base Case Value', 'C'), (33, 'Bull Case Value', 'D')]:
    ws_ss.cell(row=r, column=1, value=lbl).font = Font(name='Arial')
    c = ws_ss.cell(row=r, column=2)
    c.value = f"={src_col}6"; c.font = BLACK_FONT; c.number_format = PRICE_FMT

# ── PROFESSOR'S NOTES ──
ws_ss.cell(row=35, column=1, value="PROFESSOR'S NOTES").font = BOLD_FONT

notes_text = [
    (36, f"Bull: {assumptions['scenarios']['Bull']['narrative']}"),
    (37, f"  NOPAT margins: {assumptions['scenarios']['Bull']['margin_vector'][0]:.0%} → {assumptions['scenarios']['Bull']['margin_vector'][-1]:.0%}. Terminal ROE > CoE → persistent advantage."),
    (39, f"Base: {assumptions['scenarios']['Base']['narrative']}"),
    (40, f"  NOPAT margins: {assumptions['scenarios']['Base']['margin_vector'][0]:.0%} → {assumptions['scenarios']['Base']['margin_vector'][-1]:.0%}. Moderate deceleration."),
    (42, f"Bear: {assumptions['scenarios']['Bear']['narrative']}"),
    (43, f"  NOPAT margins: {assumptions['scenarios']['Bear']['margin_vector'][0]:.0%} → {assumptions['scenarios']['Bear']['margin_vector'][-1]:.0%}. Competition erodes pricing power."),
    (45, 'Key sensitivities: beta/CoE, NOPAT margin trajectory, terminal growth rate.'),
    (46, f"Note: BV/share is very low (~$2.60) due to NVIDIA's asset-light model and massive buybacks."),
    (47, 'Residual income models inherently struggle with high P/B companies (NVDA P/B ~73x).'),
]
for r, txt in notes_text:
    ws_ss.cell(row=r, column=1, value=txt).font = Font(name='Arial', italic=True, color='666666')

# ── SAVE ──
OUTPUT = '/sessions/zen-zealous-hamilton/mnt/BAVGems/NVDA_Integrated_Financials.xlsx'
wb.save(OUTPUT)
print(f"Saved to {OUTPUT}")
print(f"Tabs: {wb.sheetnames}")
