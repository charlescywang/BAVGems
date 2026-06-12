#!/usr/bin/env python3
"""Generate the ZENW (Zenwave Systems, fictional) eval fixtures.

Outputs, written next to this script:
  ZENW_source_workbook.xlsx    — IS/BS/CF source tabs only (Stage 2 output state).
                                 Input for the Stage-3 / Gate-B eval.
  ZENW_analysis_workbook.xlsx  — source tabs + Condensed Financials + ALT DuPont
                                 as cached values (post-Stage-3 state). Input for
                                 the Stage-4 / assumption-map / Gate-C evals.
  ZENW_assumptions.json        — canonical schema v2, results: null (Stage-1
                                 output state; Stage-4 eval input).
  zenw_update_vault_assumptions.json — same, with `results` computed by the
                                 residual-income math below (used by the clean-v2
                                 vault fixture in bav-update's evals).
  zenw_fixture_report.txt      — consistency report: checksums, segment blend,
                                 terminal RNOA calibration, computed IVPS.

ZENW is fictional so the evals are hermetic (no EDGAR, no web) and no real
company's data is fabricated. Every checksum the skills enforce is asserted
here, so a failing eval points at the skill, never the fixture.
"""

import json
from datetime import date
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
YEARS = [2021, 2022, 2023, 2024, 2025]
FY_END = [date(y, 12, 31) for y in YEARS]
NUM_FMT = "#,##0;(#,##0)"
PCT_FMT = "0.0%"

# ---------------------------------------------------------------- financials
REV = [6000, 7200, 8500, 9900, 11400]

def pct(p):
    return [round(r * p) for r in REV]

IS_LINES = []  # (label, values, bold)
COGS = [-v for v in pct(0.40)]
RND = [-v for v in pct(0.18)]
SNM = [-v for v in pct(0.15)]
GNA = [-v for v in pct(0.07)]
GP = [REV[i] + COGS[i] for i in range(5)]
OI = [GP[i] + RND[i] + SNM[i] + GNA[i] for i in range(5)]
INT_EXP = [-100] * 5
INT_INC = [150, 160, 170, 180, 190]
OTHER = [5] * 5
PRETAX = [OI[i] + INT_EXP[i] + INT_INC[i] + OTHER[i] for i in range(5)]
TAX = [-round(0.18 * p) for p in PRETAX]
NI = [PRETAX[i] + TAX[i] for i in range(5)]

IS_LINES = [
    ("Revenue", REV, False),
    ("Cost of revenue", COGS, False),
    ("Gross profit", GP, True),
    ("Research and development", RND, False),
    ("Sales and marketing", SNM, False),
    ("General and administrative", GNA, False),
    ("Operating income", OI, True),
    ("Interest expense", INT_EXP, False),
    ("Interest income", INT_INC, False),
    ("Other income (expense), net", OTHER, False),
    ("Income before income taxes", PRETAX, True),
    ("Provision for income taxes", TAX, False),
    ("Net income", NI, True),
]

# Balance sheet. Classification key (Stage-3 vocabulary):
#   OWCA / OWCL / OLTA / OLTL / FA / FL ; AMBIG marks Gate-B judgment items.
BS_ASSETS = [
    ("Cash and cash equivalents", [1800, 2025, 2250, 2550, 2850], "FA", False),
    ("Short-term investments", [1200, 1350, 1500, 1650, 1800], "FA", True),
    ("Accounts receivable, net", pct(0.15), "OWCA", False),
    ("Inventories", pct(0.08), "OWCA", False),
    ("Prepaid expenses and other current assets", pct(0.03), "OWCA", False),
    ("Property and equipment, net", pct(0.25), "OLTA", False),
    ("Operating lease right-of-use assets", pct(0.06), "OLTA", True),
    ("Goodwill", [700] * 5, "OLTA", False),
    ("Acquired intangible assets, net", [250, 230, 210, 190, 170], "OLTA", False),
    ("Deferred income tax assets", [150, 170, 190, 210, 230], "OLTA", True),
    ("Investments in equity-method affiliates", [200, 210, 220, 230, 240], "OLTA", True),
    ("Other non-current assets", pct(0.02), "OLTA", False),
]
BS_LIABS = [
    ("Accounts payable", pct(0.07), "OWCL", False),
    ("Accrued compensation and benefits", pct(0.05), "OWCL", False),
    ("Deferred revenue, current", pct(0.10), "OWCL", False),
    ("Operating lease liabilities, current", [72, 86, 102, 119, 137], "OWCL", True),
    ("Other current liabilities", [108, 130, 153, 178, 205], "OWCL", False),
    ("Long-term debt", [2000] * 5, "FL", False),
    ("Operating lease liabilities, non-current", [300, 360, 425, 495, 570], "OLTL", True),
    ("Deferred income tax liabilities", [100, 115, 130, 145, 160], "OLTL", True),
    ("Pension and other postretirement obligations", [250, 255, 260, 265, 270], "OLTL", True),
    ("Other long-term liabilities", [90, 108, 127, 149, 171], "OLTL", False),
]
CURRENT_ASSET_LABELS = {l for l, _, _, _ in BS_ASSETS[:5]}
CURRENT_LIAB_LABELS = {l for l, _, _, _ in BS_LIABS[:5]}

CS_APIC = [1400, 1500, 1600, 1700, 1800]
AOCI = [-50, -60, -55, -45, -40]

TOTAL_ASSETS = [sum(v[i] for _, v, _, _ in BS_ASSETS) for i in range(5)]
TOTAL_LIABS = [sum(v[i] for _, v, _, _ in BS_LIABS) for i in range(5)]
RETAINED = [TOTAL_ASSETS[i] - TOTAL_LIABS[i] - CS_APIC[i] - AOCI[i] for i in range(5)]
TOTAL_EQUITY = [CS_APIC[i] + RETAINED[i] + AOCI[i] for i in range(5)]
for i in range(5):
    assert TOTAL_ASSETS[i] == TOTAL_LIABS[i] + TOTAL_EQUITY[i], "BS must balance"

# Cash flow — CFF buyback plug ties net change to the BS cash line.
CASH = [v for l, v, _, _ in BS_ASSETS if l == "Cash and cash equivalents"][0]
BEGIN_CASH = [1500] + CASH[:-1]
DNA = [300, 360, 425, 495, 570]
SBC = [200, 240, 283, 330, 380]
DEF_TAX_CF = [-10, -5, -5, -5, -5]
CFO_TARGET = [1400, 1650, 1950, 2300, 2650]
WC_CHANGE = [CFO_TARGET[i] - NI[i] - DNA[i] - SBC[i] - DEF_TAX_CF[i] for i in range(5)]
CAPEX = [-550, -640, -730, -825, -925]
INV_PURCH = [-400, -450, -500, -550, -600]
INV_SALES = [300, 350, 400, 450, 500]
CFI = [CAPEX[i] + INV_PURCH[i] + INV_SALES[i] for i in range(5)]
STOCK_ISSUED = [60, 70, 80, 90, 100]
NET_CHANGE = [CASH[i] - BEGIN_CASH[i] for i in range(5)]
CFF = [NET_CHANGE[i] - CFO_TARGET[i] - CFI[i] for i in range(5)]
BUYBACK = [CFF[i] - STOCK_ISSUED[i] for i in range(5)]

CF_LINES = [
    ("Net income", NI, False),
    ("Depreciation and amortization", DNA, False),
    ("Stock-based compensation", SBC, False),
    ("Deferred income taxes", DEF_TAX_CF, False),
    ("Changes in operating assets and liabilities", WC_CHANGE, False),
    ("Net cash provided by operating activities", CFO_TARGET, True),
    ("Purchases of property and equipment", CAPEX, False),
    ("Purchases of investments", INV_PURCH, False),
    ("Maturities and sales of investments", INV_SALES, False),
    ("Net cash used in investing activities", CFI, True),
    ("Repurchases of common stock", BUYBACK, False),
    ("Proceeds from issuance of common stock", STOCK_ISSUED, False),
    ("Net cash used in financing activities", CFF, True),
    ("Effect of exchange rate changes on cash", [0] * 5, False),
    ("Net change in cash and cash equivalents", NET_CHANGE, True),
    ("Cash and cash equivalents, beginning of period", BEGIN_CASH, False),
    ("Cash and cash equivalents, end of period", CASH, True),
]
for i in range(5):
    assert CFO_TARGET[i] + CFI[i] + CFF[i] == NET_CHANGE[i]
    assert BEGIN_CASH[i] + NET_CHANGE[i] == CASH[i]

# ------------------------------------------------- condensed / DuPont values
CAT_NAMES = {
    "OWCA": "Op. Working Capital Asset",
    "OWCL": "Op. Working Capital Liability",
    "OLTA": "Op. Long-Term Asset",
    "OLTL": "Op. Long-Term Liability",
    "FA": "Financial Asset",
    "FL": "Financial Liability",
}

def agg(cat):
    rows = [v for _, v, c, _ in BS_ASSETS + BS_LIABS if c == cat]
    return [sum(r[i] for r in rows) for i in range(5)]

OWCA, OWCL, OLTA, OLTL = agg("OWCA"), agg("OWCL"), agg("OLTA"), agg("OLTL")
FIN_A, FIN_L = agg("FA"), agg("FL")
NOWC = [OWCA[i] - OWCL[i] for i in range(5)]
NET_OP_LTA = [OLTA[i] - OLTL[i] for i in range(5)]
NOA = [NOWC[i] + NET_OP_LTA[i] for i in range(5)]
NET_DEBT = [FIN_L[i] - FIN_A[i] for i in range(5)]
EQ_IMPLIED = [NOA[i] - NET_DEBT[i] for i in range(5)]
TOTAL_CAPITAL = [NET_DEBT[i] + EQ_IMPLIED[i] for i in range(5)]
for i in range(5):
    assert EQ_IMPLIED[i] == TOTAL_EQUITY[i], "implied equity must equal reported"
    assert TOTAL_CAPITAL[i] == NOA[i], "NOA = Total Capital must hold"

ETR = [-TAX[i] / PRETAX[i] for i in range(5)]
NET_INT = [-INT_EXP[i] - INT_INC[i] for i in range(5)]  # expense positive
NIAT = [NET_INT[i] * (1 - ETR[i]) for i in range(5)]
NOPAT = [NI[i] + NIAT[i] for i in range(5)]

def avg(series, i):
    return (series[i] + series[i - 1]) / 2

DUPONT = {l: ["n/a"] for l in [
    "Sales Growth", "NOPAT Margin", "Avg NOA", "Asset Turnover", "RNOA",
    "Avg Net Debt", "After-tax CoD", "Spread", "Avg Equity", "FLEV",
    "Leverage Gain", "ROE (decomposed)", "Actual ROE", "Check"]}
DUPONT["NOPAT Margin"] = [NOPAT[0] / REV[0]]
for i in range(1, 5):
    rnoa = NOPAT[i] / avg(NOA, i)
    cod = NIAT[i] / avg(NET_DEBT, i)
    flev = avg(NET_DEBT, i) / avg(EQ_IMPLIED, i)
    decomposed = rnoa + flev * (rnoa - cod)
    actual = NI[i] / avg(EQ_IMPLIED, i)
    assert abs(decomposed - actual) < 1e-9, "DuPont identity must reconcile"
    DUPONT["Sales Growth"].append(REV[i] / REV[i - 1] - 1)
    DUPONT["NOPAT Margin"].append(NOPAT[i] / REV[i])
    DUPONT["Avg NOA"].append(avg(NOA, i))
    DUPONT["Asset Turnover"].append(REV[i] / avg(NOA, i))
    DUPONT["RNOA"].append(rnoa)
    DUPONT["Avg Net Debt"].append(avg(NET_DEBT, i))
    DUPONT["After-tax CoD"].append(cod)
    DUPONT["Spread"].append(rnoa - cod)
    DUPONT["Avg Equity"].append(avg(EQ_IMPLIED, i))
    DUPONT["FLEV"].append(flev)
    DUPONT["Leverage Gain"].append(flev * (rnoa - cod))
    DUPONT["ROE (decomposed)"].append(decomposed)
    DUPONT["Actual ROE"].append(actual)
    DUPONT["Check"].append("OK")
HIST_AVG_COD = sum(DUPONT["After-tax CoD"][1:]) / 4

# ----------------------------------------------------------------- scenarios
RF, ERP, TAXR = 0.042, 0.05, 0.18
PRICE, PRICE_DATE, SHARES = 84.0, "2026-06-01", 460
SCEN = {
    "Bear": dict(
        probability=0.25, beta=1.30, tg=0.025,
        growth=[0.07, 0.06, 0.055, 0.05, 0.045, 0.04, 0.035, 0.03, 0.028, 0.025],
        margin=[0.15, 0.145, 0.14, 0.138, 0.135, 0.133, 0.13, 0.128, 0.126, 0.125],
        nowc=[0.012, 0.024, 0.036, 0.048, 0.06, 0.07, 0.08, 0.088, 0.095, 0.10],
        nola=[0.40, 0.455, 0.51, 0.565, 0.615, 0.665, 0.71, 0.755, 0.795, 0.835],
        target_x=1.25,
        narrative=("White-box/merchant-silicon substitution hollows out hardware, "
                   "and the observability platform loses the AIOps transition to "
                   "hyperscaler-native tooling; pricing power erodes."),
        seg_growth={"Observability Platform": [0.10, 0.09, 0.085],
                    "Network Infrastructure": [0.04, 0.03, 0.025],
                    "Professional Services & Support": [0.035, 0.03, 0.025]},
    ),
    "Base": dict(
        probability=0.50, beta=1.15, tg=0.035,
        growth=[0.13, 0.12, 0.11, 0.10, 0.09, 0.08, 0.07, 0.06, 0.045, 0.035],
        margin=[0.16, 0.16, 0.158, 0.158, 0.157, 0.157, 0.156, 0.156, 0.155, 0.155],
        nowc=[0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.085, 0.09],
        nola=[0.38, 0.42, 0.46, 0.50, 0.54, 0.575, 0.61, 0.64, 0.665, 0.69],
        target_x=2.0,
        narrative=("Platform keeps compounding on the installed hardware base; "
                   "attach rate and NRR hold; hardware grows with campus refresh "
                   "cycles; margins flat-to-slightly-down as mix shifts."),
        seg_growth={"Observability Platform": [0.18, 0.17, 0.155],
                    "Network Infrastructure": [0.09, 0.08, 0.075],
                    "Professional Services & Support": [0.085, 0.075, 0.07]},
    ),
    "Bull": dict(
        probability=0.25, beta=1.05, tg=0.04,
        growth=[0.18, 0.17, 0.155, 0.14, 0.125, 0.11, 0.095, 0.08, 0.06, 0.04],
        margin=[0.165, 0.17, 0.172, 0.174, 0.176, 0.178, 0.18, 0.18, 0.18, 0.18],
        nowc=[0.01, 0.018, 0.026, 0.034, 0.042, 0.05, 0.057, 0.063, 0.069, 0.075],
        nola=[0.37, 0.395, 0.42, 0.445, 0.47, 0.49, 0.51, 0.528, 0.545, 0.56],
        target_x=3.0,
        narrative=("AI-driven network telemetry becomes a must-have: platform ARR "
                   "accelerates, NRR >120%, software mix lifts NOPAT margin to 18% "
                   "with durable switching costs."),
        seg_growth={"Observability Platform": [0.24, 0.23, 0.21],
                    "Network Infrastructure": [0.135, 0.125, 0.11],
                    "Professional Services & Support": [0.11, 0.10, 0.095]},
    ),
}
SEG_SHARE = {"Observability Platform": 0.45,
             "Network Infrastructure": 0.35,
             "Professional Services & Support": 0.20}

for name, s in SCEN.items():
    ke = RF + s["beta"] * ERP
    s["ke"] = round(ke, 4)
    t_rnoa = s["margin"][-1] / (s["nowc"][-1] + s["nola"][-1])
    s["t_rnoa_planned"] = t_rnoa
    ratio = t_rnoa / ke
    assert abs(ratio - s["target_x"]) < 0.1, f"{name} calibration drifted: {ratio:.2f}x"
    for y in range(3):
        blend = sum(SEG_SHARE[k] * v[y] for k, v in s["seg_growth"].items())
        assert abs(blend - s["growth"][y]) <= 0.005, \
            f"{name} Y{y+1} segment blend {blend:.4f} vs {s['growth'][y]:.4f}"

# Residual-income model per the stage4 tab layout (beginning balances,
# leverage carried forward, after-tax CoD from DuPont history).
LEV = NET_DEBT[-1] / TOTAL_CAPITAL[-1]
B8 = HIST_AVG_COD  # rubric row 7/8 quirk: averages after-tax history, then *(1-t);
B8_after = B8 * (1 - TAXR)

def run_model(s):
    ke, g = s["ke"], s["tg"]
    sales, nowc_v, nola_v, nd_v, eq_v, nopat_v, ni_v, ae_v = [], [], [], [], [], [], [], []
    prev_sales = REV[-1]
    for t in range(10):
        st = prev_sales * (1 + s["growth"][t])
        sales.append(st)
        prev_sales = st
    for t in range(10):
        if t == 0:
            nowc_t, nola_t, nd_t = NOWC[-1], NET_OP_LTA[-1], NET_DEBT[-1]
        else:
            nowc_t = s["nowc"][t] * sales[t]
            nola_t = s["nola"][t] * sales[t]
            nd_t = LEV * (nowc_t + nola_t)
        noa_t = nowc_t + nola_t
        eq_t = noa_t - nd_t
        nopat_t = sales[t] * s["margin"][t]
        ni_t = nopat_t - nd_t * B8_after
        ae_t = ni_t - ke * eq_t
        nowc_v.append(nowc_t); nola_v.append(nola_t); nd_v.append(nd_t)
        eq_v.append(eq_t); nopat_v.append(nopat_t); ni_v.append(ni_t); ae_v.append(ae_t)
    pv = sum(ae_v[t] / (1 + ke) ** (t + 1) for t in range(10))
    tv = ae_v[9] * (1 + g) / (ke - g)
    pv_tv = tv / (1 + ke) ** 10
    iv = eq_v[0] + pv + pv_tv
    ivps = iv / SHARES
    term_rnoa = nopat_v[9] / (nowc_v[9] + nola_v[9])
    assert term_rnoa < 5 * ke, "fixture must not itself violate the 5x CoE rule"
    return ivps, term_rnoa

RESULTS = {}
for name, s in SCEN.items():
    ivps, term_rnoa = run_model(s)
    RESULTS[name] = {"ivps": round(ivps, 2), "terminalRNOA": round(term_rnoa, 4)}
W_IVPS = float(Decimal(str(sum(SCEN[n]["probability"] * RESULTS[n]["ivps"] for n in SCEN)))
               .quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))
GAP = round((W_IVPS - PRICE) / PRICE * 100, 1)

# -------------------------------------------------------------- assumptions
PROV = {
    "Bear": {
        "growthVector": "Y1-3 from segment build-up per Bear Hypothesis #1 (merchant-silicon share loss caps Infrastructure at +4%) and #2 (AIOps displacement holds Platform to +10%); fade to 2.5% terminal, below nominal GDP.",
        "marginVector": "Bear H1: hardware gross margin to low-30s on pricing concessions; mix cannot offset — NOPAT margin compresses 15.7% -> 12.5% (report §6, Bear).",
        "nowcRatioVector": "Channel inventory builds and slower collections unwind the deferred-revenue float; ramps 1% -> 10% of sales (report §2, working-capital risk factor).",
        "nolaRatioVector": "Required capacity/refresh CapEx persists while revenue stalls; ramp to 0.835 forces terminal RNOA to ~1.25x CoE (report §6 Bear quantification).",
    },
    "Base": {
        "growthVector": "Y1-3 from segment build-up in report §1/§6 Base path (Platform 18/17/15.5, Infra 9/8/7.5, Services 8.5/7.5/7); fades to 3.5% terminal ~ nominal GDP.",
        "marginVector": "Mix shift to software offsets hardware pressure; NOPAT margin ~16% easing to 15.5% terminal per report §6 Base margin-by-segment walk.",
        "nowcRatioVector": "Deferred-revenue float fades as enterprise mix matures; NOWC/Sales ramps 1% -> 9% (report §4 accounting-policy note on billings).",
        "nolaRatioVector": "Steady-state infrastructure intensity plus capitalized R&D facilities; ramp 0.345 -> 0.69 targets terminal RNOA 2.0x CoE (report §6 Base).",
    },
    "Bull": {
        "growthVector": "Y1-3 from Bull Hypothesis #1 (AI telemetry attach lifts Platform to +24%) and #2 (NRR >120% sustained); fade to 4% terminal cap per rubric.",
        "marginVector": "Software mix >55% of revenue by Y5 lifts NOPAT margin to 18% and holds — Bull H1 quantification in report §6.",
        "nowcRatioVector": "Prepaid multi-year platform deals keep NOWC lean; ramps 1% -> 7.5% (report §6 Bull).",
        "nolaRatioVector": "Asset-light software growth; ramp to 0.56 yields terminal RNOA ~3.0x CoE — durable-moat band (report §6 Bull, moat §3).",
    },
}

def assumptions_dict(with_results: bool, analysis_date: str, trigger: str):
    d = {
        "schemaVersion": 2,
        "ticker": "ZENW",
        "company": "Zenwave Systems, Inc.",
        "meta": {
            "analysisDate": analysis_date,
            "trigger": trigger,
            "supersedes": None,
            "anchorFY": "FY2025",
            "anchorFYEnd": "2025-12-31",
            "anchorRevenue": REV[-1],
            "currency": "USD millions",
            "fiscalYearEnd": "December 31",
            "dataYears": [f"FY{y}" for y in YEARS],
        },
        "marketData": {
            "price": PRICE,
            "priceDate": PRICE_DATE,
            "dilutedShares": SHARES,
            "riskFreeRate": RF,
            "equityRiskPremium": ERP,
        },
        "segments": [
            {
                "name": seg,
                "shareOfRevenue": SEG_SHARE[seg],
                "growthY1to3": {n: SCEN[n]["seg_growth"][seg] for n in ("Bear", "Base", "Bull")},
                "rationale": rationale,
            }
            for seg, rationale in [
                ("Observability Platform",
                 "ARR engine (~45% of revenue, NRR 115%); AI telemetry attach is the swing factor."),
                ("Network Infrastructure",
                 "Hardware (~35%); campus refresh cycles vs. merchant-silicon substitution."),
                ("Professional Services & Support",
                 "Drag-along (~20%); tracks the installed base with a one-year lag."),
            ]
        ],
        "kpiWatchlist": [
            {"metric": "Platform ARR growth (YoY)", "current": "19%", "asOf": "Q4 FY2025",
             "bearTrigger": "<12% for 2 consecutive quarters", "bullTrigger": ">22% sustained",
             "linkedAssumption": "growthVector Y1-3 (all scenarios)", "source": "earnings supplement"},
            {"metric": "Net revenue retention", "current": "115%", "asOf": "Q4 FY2025",
             "bearTrigger": "<108%", "bullTrigger": ">120% for 2 quarters",
             "linkedAssumption": "Base/Bull growthVector Y1-3", "source": "10-K key metrics"},
            {"metric": "Hardware segment gross margin", "current": "38%", "asOf": "Q4 FY2025",
             "bearTrigger": "<33% for 2 consecutive quarters", "bullTrigger": ">41%",
             "linkedAssumption": "marginVector (Bear compression path)", "source": "10-Q segment note"},
            {"metric": "Remaining performance obligations growth (YoY)", "current": "21%", "asOf": "Q4 FY2025",
             "bearTrigger": "<10%", "bullTrigger": ">25%",
             "linkedAssumption": "growthVector Y2-3", "source": "10-Q revenue note"},
        ],
        "scenarios": {},
        "results": None,
    }
    for n in ("Bear", "Base", "Bull"):
        s = SCEN[n]
        d["scenarios"][n] = {
            "probability": s["probability"],
            "narrative": s["narrative"],
            "beta": s["beta"],
            "costOfEquity": s["ke"],
            "taxRate": TAXR,
            "terminalGrowth": s["tg"],
            "growthVector": s["growth"],
            "marginVector": s["margin"],
            "nowcRatioVector": s["nowc"],
            "nolaRatioVector": s["nola"],
            "targetTerminalRNOAxCoE": s["target_x"],
            "provenance": PROV[n],
        }
    if with_results:
        d["results"] = {
            "asOf": analysis_date,
            "Bear": RESULTS["Bear"],
            "Base": RESULTS["Base"],
            "Bull": RESULTS["Bull"],
            "weightedIVPS": W_IVPS,
            "gapVsPricePct": GAP,
        }
    return d

# ------------------------------------------------------------------ workbooks
def header_block(ws, statement):
    ws["A1"] = "Company: Zenwave Systems, Inc. (ZENW)"
    ws["A2"] = f"Statement: {statement}"
    ws["A3"] = "Units: USD in Millions"
    ws["A4"] = "Source: ZENW eval fixture (synthetic; generated by make_zenw_fixtures.py)"
    ws["A6"] = "Line Item"
    ws["A6"].font = Font(bold=True)
    for j, d in enumerate(FY_END):
        c = ws.cell(row=6, column=2 + j, value=d)
        c.number_format = "mmm dd, yyyy"
        c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 48
    for j in range(len(FY_END)):
        ws.column_dimensions[get_column_letter(2 + j)].width = 16

def fill_rows(ws, rows, start=7, fmt=NUM_FMT):
    r = start
    for label, values, bold in rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
        for j, v in enumerate(values):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.number_format = fmt
            if bold:
                c.font = Font(bold=True)
        r += 1
    return r

def build_source_tabs(wb):
    ws = wb.active
    ws.title = "Income Statement"
    header_block(ws, "Income Statement")
    fill_rows(ws, IS_LINES)

    ws = wb.create_sheet("Balance Sheet")
    header_block(ws, "Balance Sheet")
    rows = []
    rows += [(l, v, False) for l, v, _, _ in BS_ASSETS[:5]]
    rows.append(("Total current assets",
                 [sum(v[i] for l, v, _, _ in BS_ASSETS if l in CURRENT_ASSET_LABELS) for i in range(5)], True))
    rows += [(l, v, False) for l, v, _, _ in BS_ASSETS[5:]]
    rows.append(("Total assets", TOTAL_ASSETS, True))
    rows += [(l, v, False) for l, v, _, _ in BS_LIABS[:5]]
    rows.append(("Total current liabilities",
                 [sum(v[i] for l, v, _, _ in BS_LIABS if l in CURRENT_LIAB_LABELS) for i in range(5)], True))
    rows += [(l, v, False) for l, v, _, _ in BS_LIABS[5:]]
    rows.append(("Total liabilities", TOTAL_LIABS, True))
    rows.append(("Common stock and additional paid-in capital", CS_APIC, False))
    rows.append(("Retained earnings", RETAINED, False))
    rows.append(("Accumulated other comprehensive income (loss)", AOCI, False))
    rows.append(("Total stockholders' equity", TOTAL_EQUITY, True))
    rows.append(("Total liabilities and stockholders' equity",
                 [TOTAL_LIABS[i] + TOTAL_EQUITY[i] for i in range(5)], True))
    fill_rows(ws, rows)

    ws = wb.create_sheet("Cash Flow Statement")
    header_block(ws, "Cash Flow Statement")
    fill_rows(ws, CF_LINES)

def build_analysis_tabs(wb):
    ws = wb.create_sheet("Condensed Financials")
    ws["A1"] = "Zenwave Systems, Inc. (ZENW) — Condensed Financials"
    ws["A2"] = "Fixture note: post-Stage-3 state; values cached (classifications fixed as shown)."
    ws["A4"] = "CONDENSED INCOME STATEMENT"
    ws["A4"].font = Font(bold=True)
    for j, d in enumerate(FY_END):
        c = ws.cell(row=5, column=2 + j, value=d)
        c.number_format = "mmm dd, yyyy"
        c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 48
    is_rows = [
        ("Net Income", NI, False),
        ("Interest Expense", [-v for v in INT_EXP], False),
        ("Interest Income", INT_INC, False),
        ("Net Interest Expense", NET_INT, False),
        ("Pretax Income", PRETAX, False),
        ("Tax Expense", [-t for t in TAX], False),
        ("Effective Tax Rate", ETR, False),
        ("Net Interest After Tax", NIAT, False),
        ("NOPAT", NOPAT, True),
    ]
    r = 6
    for label, values, bold in is_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
        for j, v in enumerate(values):
            c = ws.cell(row=r, column=2 + j, value=v)
            c.number_format = PCT_FMT if label == "Effective Tax Rate" else "#,##0.0;(#,##0.0)"
            if bold:
                c.font = Font(bold=True)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="BALANCE SHEET CLASSIFICATION").font = Font(bold=True)
    ws.cell(row=r, column=2, value="Classification")
    r += 1
    for label, values, cat, ambig in BS_ASSETS + BS_LIABS:
        ws.cell(row=r, column=1, value=label)
        tag = CAT_NAMES[cat] + (" ⚠️ (User: confirmed at Gate B)" if ambig else "")
        ws.cell(row=r, column=2, value=tag)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="CONDENSED BALANCE SHEET").font = Font(bold=True)
    for j, d in enumerate(FY_END):
        c = ws.cell(row=r, column=2 + j, value=d)
        c.number_format = "mmm dd, yyyy"
        c.font = Font(bold=True)
    r += 1
    bs_rows = [
        ("Operating WC Assets", OWCA, False),
        ("Operating WC Liabilities", OWCL, False),
        ("NOWC", NOWC, True),
        ("Operating LT Assets", OLTA, False),
        ("Operating LT Liabilities", OLTL, False),
        ("Net Operating LT Assets", NET_OP_LTA, True),
        ("NOA", NOA, True),
        ("Financial Liabilities", FIN_L, False),
        ("Financial Assets", FIN_A, False),
        ("Net Debt", NET_DEBT, True),
        ("Equity (implied)", EQ_IMPLIED, False),
        ("Reported Equity", TOTAL_EQUITY, False),
        ("Unclassified residual", [0] * 5, False),
        ("Total Capital", TOTAL_CAPITAL, True),
        ("Check", ["OK"] * 5, True),
    ]
    for label, values, bold in bs_rows:
        ws.cell(row=r, column=1, value=label).font = Font(bold=bold)
        for j, v in enumerate(values):
            c = ws.cell(row=r, column=2 + j, value=v)
            if isinstance(v, (int, float)):
                c.number_format = NUM_FMT
            if bold:
                c.font = Font(bold=True)
        r += 1

    ws = wb.create_sheet("ALT DuPont")
    ws["A1"] = "Zenwave Systems, Inc. (ZENW) — ALT DuPont (ROE = RNOA + FLEV × Spread)"
    ws["A2"] = ("Note: Zenwave Systems has negative Net Debt — FLEV is negative, "
                "so leverage REDUCES ROE relative to RNOA.")
    for j, d in enumerate(FY_END):
        c = ws.cell(row=4, column=2 + j, value=d)
        c.number_format = "mmm dd, yyyy"
        c.font = Font(bold=True)
    ws.column_dimensions["A"].width = 48
    r = 5
    pct_rows = {"Sales Growth", "NOPAT Margin", "RNOA", "After-tax CoD", "Spread",
                "Leverage Gain", "ROE (decomposed)", "Actual ROE"}
    for label in ["Sales Growth", "NOPAT Margin", "Avg NOA", "Asset Turnover", "RNOA",
                  "Avg Net Debt", "After-tax CoD", "Spread", "Avg Equity", "FLEV",
                  "Leverage Gain", "ROE (decomposed)", "Actual ROE", "Check"]:
        ws.cell(row=r, column=1, value=label)
        series = DUPONT[label]
        for j in range(5):
            v = series[j] if j < len(series) else None
            c = ws.cell(row=r, column=2 + j, value=v)
            if isinstance(v, float):
                c.number_format = PCT_FMT if label in pct_rows else "#,##0.00;(#,##0.00)"
        r += 1

# --------------------------------------------------------------------- main
def main():
    wb = Workbook()
    build_source_tabs(wb)
    wb.save(HERE / "ZENW_source_workbook.xlsx")

    wb = Workbook()
    build_source_tabs(wb)
    build_analysis_tabs(wb)
    wb.save(HERE / "ZENW_analysis_workbook.xlsx")

    (HERE / "ZENW_assumptions.json").write_text(
        json.dumps(assumptions_dict(False, "2026-06-01", "initial_build"), indent=2) + "\n")
    (HERE / "zenw_update_vault_assumptions.json").write_text(
        json.dumps(assumptions_dict(True, "2026-06-01", "initial_build"), indent=2) + "\n")

    lines = ["ZENW fixture consistency report", "=" * 40]
    lines.append(f"FY2025 anchor: revenue {REV[-1]:,}  NOPAT {NOPAT[-1]:,.1f}  "
                 f"margin {NOPAT[-1]/REV[-1]:.3%}")
    lines.append(f"FY2025 NOWC {NOWC[-1]:,}  NOLA {NET_OP_LTA[-1]:,}  NOA {NOA[-1]:,}  "
                 f"NetDebt {NET_DEBT[-1]:,}  Equity {EQ_IMPLIED[-1]:,}")
    lines.append(f"NOWC/Sales {NOWC[-1]/REV[-1]:.3f}   NOLA/Sales {NET_OP_LTA[-1]/REV[-1]:.3f}   "
                 f"leverage (ND/TC) {LEV:.3f}")
    lines.append(f"Hist avg after-tax CoD (DuPont) {HIST_AVG_COD:.4f}; model B8 {B8_after:.4f}")
    lines.append("BS balances, CF ties, NOA=Total Capital, DuPont reconciles: all asserted OK")
    for n in ("Bear", "Base", "Bull"):
        s = SCEN[n]
        lines.append(f"{n}: Ke {s['ke']:.4f}  terminal RNOA {RESULTS[n]['terminalRNOA']:.3f} "
                     f"({RESULTS[n]['terminalRNOA']/s['ke']:.2f}x CoE, target {s['target_x']}x)  "
                     f"IVPS {RESULTS[n]['ivps']}")
    lines.append(f"Weighted IVPS {W_IVPS} vs price {PRICE} -> gap {GAP}%")
    (HERE / "zenw_fixture_report.txt").write_text("\n".join(lines) + "\n")
    print("\n".join(lines))

if __name__ == "__main__":
    main()
